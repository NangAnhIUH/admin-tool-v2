import streamlit as st
import json
import pandas as pd
import io
import os
import subprocess
import re
import shutil
import zipfile
import tempfile
import xlsxwriter
import openpyxl
from datetime import datetime
import streamlit.components.v1 as components
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from email.mime.base import MIMEBase
from email import encoders

# ==========================================
# HELPER FUNCTIONS FOR EMAIL
# ==========================================
def normalize_func_name(name):
    """Chuẩn hóa tên function để so sánh."""
    if not isinstance(name, str): return ""
    name = name.replace('_', ' ').replace('-', ' ')
    name = name.upper()
    return " ".join(name.split())

def extract_name_from_email(email):
    """Lấy tên người nhận từ email."""
    try:
        if '@' in email:
            local_part = email.split('@')[0]
            return local_part.split('.')[0].capitalize() if '.' in local_part else local_part.capitalize()
    except: pass
    return "Partner"

def get_email_mapping_from_upload(file_obj):
    """Đọc mapping email từ file upload (Sheet 'Aprrove')."""
    try:
        file_obj.seek(0)
        # Read header=1 to get column names like 'Funtional', 'Functional admin'
        df = pd.read_excel(file_obj, sheet_name='Aprrove', header=1)
        df.columns = [str(col).strip() for col in df.columns]
        
        # Identify Key Column (Function Name)
        # User instruction: "cột đầu vào đó chính là cột đầu tiên của APPROVAL_FILE"
        # In the file structure, 'STT' is col 0, 'Funtional' is col 1.
        # We will look for 'Funtional'/'Functional' OR strictly use index 1 if available.
        
        func_col = next((c for c in df.columns if 'Funtional' in c or 'Functional' in c), None)
        if not func_col and len(df.columns) > 1:
             func_col = df.columns[1] # Fallback to 2nd column (index 1)
             
        to_email_col = next((c for c in df.columns if 'Functional admin' in c), None)
        
        if not func_col or not to_email_col: return {}
        
        # Create dictionary: Normalized_Key -> {raw_name, to, cc}
        email_map = {}
        for _, row in df.iterrows():
            func_name = row[func_col]
            to_email = row[to_email_col]
            
            if pd.notna(func_name):
                # Use raw name for display, normalized for lookup
                raw_key = str(func_name).strip()
                norm_key = normalize_func_name(raw_key)
                
                if not norm_key: continue
                
                current = email_map.get(norm_key, {'name': raw_key, 'to': '', 'cc': ''})
                
                if pd.notna(to_email):
                    val = str(to_email).strip()
                    if val and val.lower() != 'nan': current['to'] = val
                
                current['cc'] = ""
                email_map[norm_key] = current
                
        return email_map
    except Exception as e:
        st.error(f"Lỗi đọc file Approval (Sheet 'Aprrove'): {e}")
        return {}

def create_eml_draft(output_root_dir):
    """
    Quét thư mục Output, tìm file Excel và thông tin Email để tạo file .eml
    """
    logs = []
    count_success = 0
    
    if not os.path.exists(output_root_dir):
        return [f"❌ Thư mục gốc không tồn tại: {output_root_dir}"]

    # Duyệt qua các thư mục con (Group Function)
    for group_name in os.listdir(output_root_dir):
        group_path = os.path.join(output_root_dir, group_name)
        if not os.path.isdir(group_path): continue

        # Tìm các file cần thiết trong thư mục nhóm
        files = os.listdir(group_path)
        
        # 1. Tìm file HTML Email (để lấy nội dung & người nhận)
        # Format: email_{group_name}_{recipient_email}.html
        html_file = None
        current_email = ""
        
        for f in files:
            if f.startswith("email_") and f.endswith(".html"):
                html_file = f
                # Extract email logic: email_{group}_{email}.html
                temp = f.replace("email_", "").replace(".html", "")
                
                # Heuristic: Find the separator before the email.
                # Email usually contains '@'.
                if '@' in temp:
                    at_index = temp.find('@')
                    # Find last underscore before at_index
                    sep_index = temp.rfind('_', 0, at_index)
                    if sep_index != -1:
                        current_email = temp[sep_index+1:]
                    else:
                        # Fallback: maybe the whole temp is the email (unlikely if group exists)
                        # Or maybe underscore replacement for special chars mess this up.
                        pass
                break

        # 2. Tìm file Excel Bảng Kê (để đính kèm)
        excel_file = next((f for f in files if f.endswith('.xlsx') and f.startswith('BK_GRAB_')), None)

        if html_file and excel_file:
            try:
                # Đọc nội dung HTML
                html_path = os.path.join(group_path, html_file)
                with open(html_path, 'r', encoding='utf-8') as f:
                    email_body = f.read()
                
                to_email = current_email if current_email else ""
                
                # Tạo đối tượng Email (Structure: Mixed -> [Alternative(Text, HTML), Attachments])
                msg = MIMEMultipart('mixed')
                msg['Subject'] = f"Bảng Kê Grab Tháng - Nhóm {group_name}"
                if to_email: msg['To'] = to_email
                
                # Phần nội dung (Body)
                msg_body = MIMEMultipart('alternative')
                text_part = MIMEText("Vui lòng xem nội dung HTML đính kèm hoặc mở bằng trình duyệt hỗ trợ.", 'plain', 'utf-8')
                html_part = MIMEText(email_body, 'html', 'utf-8')
                msg_body.attach(text_part)
                msg_body.attach(html_part)
                msg.attach(msg_body)
                
                # Đính kèm Excel
                excel_path = os.path.join(group_path, excel_file)
                with open(excel_path, 'rb') as f:
                    part = MIMEApplication(f.read(), Name=excel_file)
                part['Content-Disposition'] = f'attachment; filename="{excel_file}"'
                msg.attach(part)

                # 3. Tạo Zip file chứa PDF
                zip_filename = f"Invoices_{group_name}.zip"
                zip_path = os.path.join(group_path, zip_filename)
                
                # Xóa file zip cũ nếu có để tránh lỗi
                if os.path.exists(zip_path):
                    try: os.remove(zip_path)
                    except: pass

                pdf_files_to_zip = []
                for root, dirs, files_in_dir in os.walk(group_path):
                     for file in files_in_dir:
                        if file.lower().endswith('.pdf'):
                            pdf_files_to_zip.append(os.path.join(root, file))
                
                pdf_count = len(pdf_files_to_zip)
                
                if pdf_count > 0:
                    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                        for pdf_p in pdf_files_to_zip:
                             # Calculate relative path (e.g. HoaDon CK/file.pdf)
                             rel_path = os.path.relpath(pdf_p, group_path)
                             zipf.write(pdf_p, arcname=rel_path)
                    
                    # Đính kèm Zip
                    if os.path.exists(zip_path):
                         with open(zip_path, 'rb') as f:
                            part = MIMEApplication(f.read(), Name=zip_filename)
                         part['Content-Disposition'] = f'attachment; filename="{zip_filename}"'
                         msg.attach(part)
                else:
                    logs.append(f"⚠️ Nhóm {group_name}: Không tìm thấy file PDF nào để tạo Zip.")
                
                # Lưu file .eml
                eml_name = f"Draft_{group_name}.eml"
                eml_path = os.path.join(group_path, eml_name)
                
                with open(eml_path, 'wb') as f:
                    f.write(msg.as_bytes())
                
                count_success += 1
                msg_log = f"✅ Đã tạo draft: {eml_name} (Email: {to_email}"
                if pdf_count > 0:
                    msg_log += f", Zip PDF: {pdf_count} file)"
                else:
                    msg_log += ", ⚠️ Không có PDF)"
                logs.append(msg_log)
                
            except Exception as e:
                logs.append(f"❌ Lỗi nhóm {group_name}: {e}")
        else:
            pass
            
    if count_success == 0:
        logs.append("⚠️ Không tạo được file nào. Vui lòng kiểm tra lại xem đã chạy 'Phân Phối PDF' chưa?")
    else:
        logs.append(f"🎉 Hoàn tất! Đã tạo {count_success} file email nháp.")
        
    return logs

# ==========================================
# 1. CẤU HÌNH & HẰNG SỐ
# ==========================================

CONFIG_FILE = 'config.json'

def load_config():
    try:
        with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        return {} 

def save_config(config):
    # Ensure directory exists if needed, though usually it's in root
    try:
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(config, f, indent=4, ensure_ascii=False)
        # Update global CONFIG to match saved state
        global CONFIG
        CONFIG = config
        # Reload constants might be needed if they are used globally, 
        # but in Streamlit, rerunning the script (which st.experimental_rerun or manual rerun does)
        # will reload constants from the new CONFIG.
    except Exception as e:
        st.error(f"Lỗi khi lưu cấu hình: {e}")

# Load config globally
CONFIG = load_config()

def get_conf(section, key, default=None):
    return CONFIG.get(section, {}).get(key, default)

# --- General ---
BANG_KE_SHEET_NAME = get_conf('general', 'BANG_KE_SHEET_NAME', 'hóa đơn chi tiết')
GROUP_FUNCTION_APPROVAL_SHEET = get_conf('general', 'GROUP_FUNCTION_APPROVAL_SHEET', 'Group')
SHEET_INTRO = "tổng quan" 

# --- Dòng bắt đầu đọc (Default values from notebook) ---
SKIPROWS_BANG_KE = get_conf('general', 'SKIPROWS_BANG_KE', 3)
SKIPROWS_EXPRESS = get_conf('general', 'SKIPROWS_EXPRESS', 7)
SKIPROWS_TRANSPORT = get_conf('general', 'SKIPROWS_TRANSPORT', 7)
SKIPROWS_GROUP_FUNCTION_APPROVAL = get_conf('general', 'SKIPROWS_GROUP_FUNCTION_APPROVAL', 1)

# --- Input Columns (Mapped to Globals for compatibility) ---
IN_COLS = CONFIG.get('input_columns', {})
IN_COL_BK_BOOKING_ID = IN_COLS.get('IN_COL_BK_BOOKING_ID', 'booking_code_for_business_grab_com')
IN_COL_BK_GROUP_NAME = IN_COLS.get('IN_COL_BK_GROUP_NAME', 'Tên nhóm (Group Name)')
IN_COL_BK_VERTICAL = IN_COLS.get('IN_COL_BK_VERTICAL', 'Dịch vụ (Vertical)')
IN_COL_BK_COMPANY_NAME = IN_COLS.get('IN_COL_BK_COMPANY_NAME', 'Tên công ty (Company Name)')
IN_COL_BK_COST_TRANS = IN_COLS.get('IN_COL_BK_COST_TRANS', 'Cước phí vận chuyển trước thuế')
IN_COL_BK_VAT_TRANS = IN_COLS.get('IN_COL_BK_VAT_TRANS', 'Thuế cước phí vận chuyển (8%)')
IN_COL_BK_COST_SERV = IN_COLS.get('IN_COL_BK_COST_SERV', 'Phí dịch vụ trước thuế')
IN_COL_BK_VAT_SERV = IN_COLS.get('IN_COL_BK_VAT_SERV', 'Thuế phí dịch vụ (8%)')
IN_COL_BK_TOTAL = IN_COLS.get('IN_COL_BK_TOTAL', 'Tổng thanh toán')

IN_COL_ET_BOOKING_ID = IN_COLS.get('IN_COL_ET_BOOKING_ID', 'Booking ID')
IN_COL_ET_EMP_ID = IN_COLS.get('IN_COL_ET_EMP_ID', 'Employee ID')
IN_COL_ET_EMP_GROUP = IN_COLS.get('IN_COL_ET_EMP_GROUP', 'Employee Group')
IN_COL_ET_EMP_NAME = IN_COLS.get('IN_COL_ET_EMP_NAME', 'Employee Name')
IN_COL_ET_TRIP_DESC = IN_COLS.get('IN_COL_ET_TRIP_DESC', 'Trip Description')
IN_COL_ET_PICKUP = IN_COLS.get('IN_COL_ET_PICKUP', 'Pick-Up Address')
IN_COL_ET_DROPOFF = IN_COLS.get('IN_COL_ET_DROPOFF', 'Drop-Off Address')
IN_COL_ET_TIME = IN_COLS.get('IN_COL_ET_TIME', 'Date & Time (GMT+7)')
IN_COL_ET_CITY = IN_COLS.get('IN_COL_ET_CITY', 'City')

IN_COL_FUNC_GROUP_PORTAL = IN_COLS.get('IN_COL_FUNC_GROUP_PORTAL', 'Group Portal')
IN_COL_FUNC_INVOICE_GROUP = IN_COLS.get('IN_COL_FUNC_INVOICE_GROUP', "Invoice's Group Function")

# --- Output Columns ---
OUT_COLS = CONFIG.get('output_columns', {})
COL_BOOKING_CODE = OUT_COLS.get('COL_BOOKING_CODE', 'Mã chuyến xe')
COL_BOOKING_CODE_ORIG = OUT_COLS.get('COL_BOOKING_CODE_ORIG', 'Mã Chuyến xe (Booking Code)')
COL_GROUP = OUT_COLS.get('COL_GROUP', 'Nhóm')
COL_EMPLOYEE_NAME = OUT_COLS.get('COL_EMPLOYEE_NAME', 'Tên nhân viên')
COL_EMPLOYEE_ID = OUT_COLS.get('COL_EMPLOYEE_ID', 'Mã nhân viên')
COL_SERVICE = OUT_COLS.get('COL_SERVICE', 'Dịch vụ')
COL_COMPANY_NAME = OUT_COLS.get('COL_COMPANY_NAME', 'Tên công ty\n(Company Name)')
COL_PAYMENT_TYPE = OUT_COLS.get('COL_PAYMENT_TYPE', 'Payment Type')
COL_CITY = OUT_COLS.get('COL_CITY', 'Tỉnh/Thành phố')

COL_COST_TRANSPORT_PRE_TAX = OUT_COLS.get('COL_COST_TRANSPORT_PRE_TAX', '(1)\nCước Phí Vận Chuyển Trước Thuế')
COL_VAT_TRANSPORT = OUT_COLS.get('COL_VAT_TRANSPORT', '(2)\nThuế Cước Phí Vận Chuyển (8%)')
COL_COST_SERVICE_PRE_TAX = OUT_COLS.get('COL_COST_SERVICE_PRE_TAX', '(3)\nPhí Dịch Vụ Trước Thuế')
COL_VAT_SERVICE = OUT_COLS.get('COL_VAT_SERVICE', '(4)\nThuế Phí Dịch Vụ (8%)')
COL_TOTAL_PRE_TAX = OUT_COLS.get('COL_TOTAL_PRE_TAX', '(5)=(1)+(3)\nTổng Tiền Trước Thuế (-Vat)')
COL_TOTAL_VAT = OUT_COLS.get('COL_TOTAL_VAT', '(6)=(2)+(4)\nTổng Tiền Thuế (Vat)')
COL_TOTAL_AMOUNT = OUT_COLS.get('COL_TOTAL_AMOUNT', '(7)=(1)+(2)+(3)+(4)\nTổng Thanh Toán (+Vat)')
COL_SURCHARGE = OUT_COLS.get('COL_SURCHARGE', 'Phụ phí không có Hóa đơn')

COL_ADJUSTED_TRIP = OUT_COLS.get('COL_ADJUSTED_TRIP', 'Chuyến xe điều chỉnh')
COL_INVOICE_NUM = OUT_COLS.get('COL_INVOICE_NUM', 'Số hóa đơn vận chuyển')
COL_LOOKUP_CODE = OUT_COLS.get('COL_LOOKUP_CODE', 'Mã tra cứu')
COL_TRIP_PURPOSE = OUT_COLS.get('COL_TRIP_PURPOSE', 'Mục đích chuyến đi')
COL_TIME = OUT_COLS.get('COL_TIME', 'Thời gian')
COL_PICKUP = OUT_COLS.get('COL_PICKUP', 'Điểm đi')
COL_DROPOFF = OUT_COLS.get('COL_DROPOFF', 'Điểm đến')
COL_LICENSE_PLATE = OUT_COLS.get('COL_LICENSE_PLATE', 'Biển kiểm soát phương tiện vận tải')
COL_GOODS_NAME = OUT_COLS.get('COL_GOODS_NAME', 'Tên hàng hóa vận chuyển')

COL_PAYMENT_METHOD_INVOICE = OUT_COLS.get('COL_PAYMENT_METHOD_INVOICE', 'Hình thức thanh toán trên Hóa đơn vận chuyển')
COL_ADJUSTED_INVOICE_NUM = OUT_COLS.get('COL_ADJUSTED_INVOICE_NUM', 'Số hóa đơn điều chỉnh')
COL_ADJUSTED_LOOKUP_CODE = OUT_COLS.get('COL_ADJUSTED_LOOKUP_CODE', 'Mã tra cứu HĐĐC')
COL_ADJUSTED_PAYMENT_METHOD = OUT_COLS.get('COL_ADJUSTED_PAYMENT_METHOD', 'Hình thức thanh toán HĐĐC')

COL_GROUP_FUNCTION = OUT_COLS.get('COL_GROUP_FUNCTION', 'Group Function')

VAL_PAYMENT_METHOD_TRANSFER = 'Chuyển khoản'
COL_ADJUSTED_INVOICE_REPORT = "Số hoá đơn"

MONEY_COL_REPORT = [
    'Tổng tiền trước thuế', 'Tổng tiền thuế', 'Tổng tiền đã có thuế',
    'Tổng tiền trước thuế 8%', 'Tổng tiền thuế 8%',
    'Tổng tiền trước thuế 10%', 'Tổng tiền thuế 10%'
]

# Money Cols for Bang Ke (Dynamic based on Output Columns)
MONEY_COLS_BANG_KE = [
    COL_COST_TRANSPORT_PRE_TAX, COL_VAT_TRANSPORT,
    COL_COST_SERVICE_PRE_TAX, COL_VAT_SERVICE,
    COL_TOTAL_PRE_TAX, COL_TOTAL_VAT, COL_TOTAL_AMOUNT
]

# --- Output Order ---
DEFAULT_OUTPUT_ORDER_KEYS = [
    'COL_BOOKING_CODE_ORIG', 'COL_BOOKING_CODE', 'COL_GROUP', 'COL_EMPLOYEE_NAME', 'COL_EMPLOYEE_ID', 'COL_SERVICE',
    'COL_COMPANY_NAME', 'COL_PAYMENT_TYPE', 
    'COL_COST_TRANSPORT_PRE_TAX', 'COL_VAT_TRANSPORT', 'COL_COST_SERVICE_PRE_TAX', 'COL_VAT_SERVICE',
    'COL_TOTAL_PRE_TAX', 'COL_TOTAL_VAT', 'COL_TOTAL_AMOUNT', 'COL_SURCHARGE',
    'COL_ADJUSTED_TRIP', 'COL_LOOKUP_CODE', 'COL_TRIP_PURPOSE', 'COL_TIME', 
    'COL_PICKUP', 'COL_DROPOFF', 'COL_LICENSE_PLATE', 'COL_INVOICE_NUM', 'COL_GOODS_NAME',
    'COL_PAYMENT_METHOD_INVOICE', 'COL_ADJUSTED_INVOICE_NUM', 'COL_ADJUSTED_LOOKUP_CODE', 
    'COL_ADJUSTED_PAYMENT_METHOD', 'COL_GROUP_FUNCTION', 'COL_CITY'
]
OUTPUT_ORDER_KEYS = CONFIG.get('output_order', DEFAULT_OUTPUT_ORDER_KEYS)

# ==========================================
# 2. HÀM HỖ TRỢ EXCEL
# ==========================================

def write_and_format_sheet_common(df, sheet_name, title_prefix, writer_obj, month, year, number_of_days, gr_func_name="ALL"):
    if df.empty: return
    df = df.dropna(axis=1, how='all')
    df = df.assign(Note=pd.Series(dtype='str'))
    df.to_excel(writer_obj, index=False, sheet_name=sheet_name, startrow=4, header=False)

    worksheet = writer_obj.sheets[sheet_name]
    workbook = writer_obj.book

    # Header Report
    title = f"{title_prefix} - NHÓM CHỨC NĂNG: {gr_func_name.upper()} - THÁNG {month} NĂM {year}"
    title_fmt = workbook.add_format({'font_size': 16, 'color': '#333333', 'bold': True})
    worksheet.write('B1', title, title_fmt)
    worksheet.write('B2', f"Từ ngày 01/{month:02d}/{year} đến ngày {number_of_days}/{month:02d}/{year}")

    # Header Table
    header_fmt = workbook.add_format({
        'align': 'center', 'valign': 'vcenter', 'text_wrap': True,
        'bold': True, 'bg_color': '#145f82', 'font_color': 'white', 'border': 1
    })
    for col_num, value in enumerate(df.columns.values):
        worksheet.write(3, col_num, value, header_fmt)

    # Column Widths & Hidden Columns
    cols_to_hide = [COL_BOOKING_CODE_ORIG, COL_COMPANY_NAME, COL_PAYMENT_TYPE]
    for idx, col in enumerate(df.columns):
        series = df[col].astype(str)
        max_len = max(series.map(len).max(), len(str(col))) + 2
        worksheet.set_column(idx, idx, min(max_len, 50))
        if col in cols_to_hide:
            worksheet.set_column(idx, idx, None, None, {'hidden': True})

    # Money Format
    money_fmt = workbook.add_format({'align': 'right', 'font_color': 'blue', 'num_format': '#,##0'})
    for col in MONEY_COLS_BANG_KE:
        if col in df.columns:
            worksheet.set_column(df.columns.get_loc(col), df.columns.get_loc(col), None, money_fmt)

    # Subtotals
    last_row_excel = 5 + len(df)
    bold_fmt = workbook.add_format({'bold': True, 'num_format': '#,##0'})
    bold_red_fmt = workbook.add_format({'bold': True, 'font_color': 'red', 'num_format': '#,##0'})
    red_fmt = workbook.add_format({'font_color': 'red'})

    if COL_SERVICE in df.columns:
        worksheet.write(2, df.columns.get_loc(COL_SERVICE), "TỔNG", bold_fmt)

    for col in MONEY_COLS_BANG_KE:
        if col in df.columns:
            col_idx = df.columns.get_loc(col)
            col_char = xlsxwriter.utility.xl_col_to_name(col_idx)
            formula = f"=SUBTOTAL(9, {col_char}5:{col_char}{last_row_excel})"
            fmt = bold_red_fmt if col == COL_TOTAL_PRE_TAX else bold_fmt
            worksheet.write_formula(2, col_idx, formula, fmt)
    
    if COL_INVOICE_NUM in df.columns:
        col_idx = df.columns.get_loc(COL_INVOICE_NUM)
        col_char = xlsxwriter.utility.xl_col_to_name(col_idx)
        worksheet.write_formula(2, col_idx, f"=SUBTOTAL(3, {col_char}5:{col_char}{last_row_excel})", red_fmt)

    # Signatures
    sig_row = last_row_excel + 2
    if COL_COST_TRANSPORT_PRE_TAX in df.columns:
        worksheet.write(sig_row, df.columns.get_loc(COL_COST_TRANSPORT_PRE_TAX), "XÁC NHẬN BẢNG KÊ", bold_fmt)
    if COL_TOTAL_AMOUNT in df.columns:
        worksheet.write(sig_row, df.columns.get_loc(COL_TOTAL_AMOUNT), "XÁC NHẬN CỦA TRƯỞNG BỘ PHẬN", bold_fmt)

def write_report_sheet(df, sheet_name, title_prefix, writer_obj):
    if df.empty: return
    df.to_excel(writer_obj, sheet_name=sheet_name, index=False, startrow=0)
    worksheet = writer_obj.sheets[sheet_name]
    workbook = writer_obj.book

    header_wrap = workbook.add_format({'bold': True, 'border': 1, 'text_wrap': True, 'valign': 'vcenter', 'bg_color': '#D9E1F2'})
    header_no_wrap = workbook.add_format({'bold': True, 'border': 1, 'text_wrap': False, 'valign': 'vcenter', 'bg_color': '#D9E1F2'})
    data_fmt = workbook.add_format({'border': 1, 'valign': 'vcenter'})
    money_fmt = workbook.add_format({'border': 1, 'num_format': '#,##0', 'valign': 'vcenter'})
    total_fmt = workbook.add_format({'bold': True, 'border': 1, 'bg_color': '#FFF2CC'})
    total_money_fmt = workbook.add_format({'bold': True, 'border': 1, 'bg_color': '#FFF2CC', 'num_format': '#,##0'})

    for col_num, value in enumerate(df.columns.values):
        fmt = header_no_wrap if value == COL_GROUP_FUNCTION else header_wrap
        worksheet.write(0, col_num, value, fmt)
        try:
            col_len = max(df.iloc[:, col_num].astype(str).map(len).max(), len(str(value))) + 2
            worksheet.set_column(col_num, col_num, min(col_len, 50))
        except:
            worksheet.set_column(col_num, col_num, 15)

    for row in range(len(df)):
        for col in range(len(df.columns)):
            val = df.iloc[row, col]
            fmt = money_fmt if df.columns[col] in MONEY_COL_REPORT else data_fmt
            if pd.isna(val): worksheet.write(row + 1, col, "", fmt)
            else: worksheet.write(row + 1, col, val, fmt)

    last_row = len(df) + 1
    worksheet.write(last_row, 0, "TỔNG CỘNG", total_fmt)
    for col_num, col_name in enumerate(df.columns):
        if col_name in MONEY_COL_REPORT:
            col_char = xlsxwriter.utility.xl_col_to_name(col_num)
            worksheet.write_formula(last_row, col_num, f"=SUM({col_char}2:{col_char}{last_row})", total_money_fmt)
        elif col_num > 0:
            worksheet.write(last_row, col_num, "", total_fmt)

def copy_intro_sheet(source_file_path, target_file_path):
    try:
        wb_source = openpyxl.load_workbook(source_file_path)
        if SHEET_INTRO not in wb_source.sheetnames: return False
        ws_source = wb_source[SHEET_INTRO]
        wb_target = openpyxl.load_workbook(target_file_path)
        if SHEET_INTRO in wb_target.sheetnames: del wb_target[SHEET_INTRO]
        ws_target = wb_target.create_sheet(SHEET_INTRO, index=0)

        for row in ws_source.rows:
            for cell in row:
                new_cell = ws_target[cell.coordinate]
                new_cell.value = cell.value
                if cell.has_style:
                    new_cell.font = cell.font.copy()
                    new_cell.border = cell.border.copy()
                    new_cell.fill = cell.fill.copy()
                    new_cell.number_format = cell.number_format
                    new_cell.alignment = cell.alignment.copy()
        
        for col_dim in ws_source.column_dimensions.values():
            if col_dim.width: ws_target.column_dimensions[col_dim.index].width = col_dim.width
        for merged in ws_source.merged_cells.ranges:
            ws_target.merge_cells(merged.coord)
            
        wb_target.save(target_file_path)
        return True
    except Exception as e:
        print(f"Error copying intro sheet: {e}")
        return False

# ==========================================
# 3. HÀM XỬ LÝ CHÍNH
# ==========================================

def process_input_data(file_bang_ke, file_express, file_transport, file_func, file_report):
    """
    Đọc và xử lý dữ liệu, trả về Dataframe kết quả để review.
    """
    # --- Đọc dữ liệu ---
    bangKe = pd.read_excel(file_bang_ke, sheet_name=BANG_KE_SHEET_NAME, skiprows=SKIPROWS_BANG_KE)
    bangKe.columns = [re.sub(r'\s+', ' ', col).strip() for col in bangKe.columns]
    
    express = pd.read_csv(file_express, skiprows=SKIPROWS_EXPRESS)
    transport = pd.read_csv(file_transport, skiprows=SKIPROWS_TRANSPORT)

    # --- Normalize Columns (Fix mismatches between Express & Transport) ---
    # Express often has 'Employee id' (lowercase) or 'Date & Time (GMT+72)'
    def normalize_columns_standard(df):
        # Strip whitespace first
        df.columns = [str(c).strip() for c in df.columns]
        
        rename_map = {}
        for col in df.columns:
            c_lower = col.lower()
            if c_lower == 'employee id':
                rename_map[col] = IN_COL_ET_EMP_ID # Standardize to Config Value (default 'Employee ID')
            elif 'date & time' in c_lower and '(gmt' in c_lower:
                # Capture inconsistent Date Time columns
                rename_map[col] = IN_COL_ET_TIME # Standardize to Config Value
        
        if rename_map:
            df = df.rename(columns=rename_map)
        return df

    express = normalize_columns_standard(express)
    transport = normalize_columns_standard(transport)

    func_df = pd.read_excel(file_func, sheet_name=GROUP_FUNCTION_APPROVAL_SHEET, skiprows=SKIPROWS_GROUP_FUNCTION_APPROVAL)
    report_df = pd.read_excel(file_report)

    # --- Mapping ---
    combined = pd.concat([express, transport], ignore_index=True)
    combined = combined.drop_duplicates(subset=[IN_COL_ET_BOOKING_ID]).set_index(IN_COL_ET_BOOKING_ID)

    bangKe[COL_EMPLOYEE_ID] = bangKe[IN_COL_BK_BOOKING_ID].map(combined[IN_COL_ET_EMP_ID])
    bangKe[IN_COL_BK_GROUP_NAME] = bangKe[IN_COL_BK_BOOKING_ID].map(combined[IN_COL_ET_EMP_GROUP])
    bangKe[COL_EMPLOYEE_NAME] = bangKe[IN_COL_BK_BOOKING_ID].map(combined[IN_COL_ET_EMP_NAME])
    bangKe[COL_TRIP_PURPOSE] = bangKe[IN_COL_BK_BOOKING_ID].map(combined[IN_COL_ET_TRIP_DESC])
    bangKe[COL_PICKUP] = bangKe[IN_COL_BK_BOOKING_ID].map(combined[IN_COL_ET_PICKUP])
    bangKe[COL_DROPOFF] = bangKe[IN_COL_BK_BOOKING_ID].map(combined[IN_COL_ET_DROPOFF])
    bangKe[COL_TIME] = bangKe[IN_COL_BK_BOOKING_ID].map(combined[IN_COL_ET_TIME])
    
    # Map City if exists
    if IN_COL_ET_CITY in combined.columns:
        bangKe[COL_CITY] = bangKe[IN_COL_BK_BOOKING_ID].map(combined[IN_COL_ET_CITY])
    else:
        bangKe[COL_CITY] = None

    # --- Regex/Keyword Extraction for City (Fallback) ---
    def extract_city_from_address(address):
        if not isinstance(address, str): return "Unknown"
        addr_lower = address.lower()
        keywords = {
            'hồ chí minh': 'Ho Chi Minh', 'ho chi minh': 'Ho Chi Minh', 'hcm': 'Ho Chi Minh',
            'hà nội': 'Hanoi', 'ha noi': 'Hanoi', 'hanoi': 'Hanoi',
            'đà nẵng': 'Da Nang', 'da nang': 'Da Nang',
            'cần thơ': 'Can Tho', 'can tho': 'Can Tho',
            'hải phòng': 'Hai Phong', 'hai phong': 'Hai Phong',
            'bình dương': 'Binh Duong', 'binh duong': 'Binh Duong',
            'đồng nai': 'Dong Nai', 'dong nai': 'Dong Nai', 'biên hòa': 'Dong Nai',
            'khánh hòa': 'Khanh Hoa', 'khanh hoa': 'Khanh Hoa', 'nha trang': 'Khanh Hoa',
            'bà rịa': 'Ba Ria - Vung Tau', 'vũng tàu': 'Ba Ria - Vung Tau', 'vung tau': 'Ba Ria - Vung Tau',
            'lâm đồng': 'Lam Dong', 'lam dong': 'Lam Dong', 'đà lạt': 'Lam Dong', 'da lat': 'Lam Dong',
            'quảng ninh': 'Quang Ninh', 'quang ninh': 'Quang Ninh', 'hạ long': 'Quang Ninh',
            'long an': 'Long An',
            'tiền giang': 'Tien Giang', 'tien giang': 'Tien Giang',
            'bắc ninh': 'Bac Ninh', 'bac ninh': 'Bac Ninh',
            'thanh hóa': 'Thanh Hoa', 'thanh hoa': 'Thanh Hoa',
            'nghệ an': 'Nghe An', 'nghe an': 'Nghe An', 'vinh': 'Nghe An',
            'thừa thiên huế': 'Hue', 'thua thien hue': 'Hue', 'huế': 'Hue',
            'quảng nam': 'Quang Nam', 'quang nam': 'Quang Nam', 'hội an': 'Quang Nam',
            'bình định': 'Binh Dinh', 'binh dinh': 'Binh Dinh', 'quy nhơn': 'Binh Dinh'
        }
        for key, val in keywords.items():
            if key in addr_lower: return val
        return None

    # Fill missing or 'Unknown' cities
    mask_need_city = bangKe[COL_CITY].isnull() | (bangKe[COL_CITY] == '') | (bangKe[COL_CITY] == 'Unknown')
    if mask_need_city.any():
        extracted_cities = bangKe.loc[mask_need_city, COL_PICKUP].apply(extract_city_from_address)
        bangKe.loc[mask_need_city, COL_CITY] = extracted_cities.fillna("Unknown")

    group_portal_map = func_df.set_index(IN_COL_FUNC_GROUP_PORTAL)[IN_COL_FUNC_INVOICE_GROUP]
    bangKe[COL_GROUP_FUNCTION] = bangKe[IN_COL_BK_GROUP_NAME].map(group_portal_map)

    # --- Đổi tên & Tính toán ---
    rename_dict = {
        IN_COL_BK_BOOKING_ID: COL_BOOKING_CODE,
        IN_COL_BK_GROUP_NAME: COL_GROUP,
        IN_COL_BK_VERTICAL: COL_SERVICE,
        IN_COL_BK_COMPANY_NAME: COL_COMPANY_NAME,
        IN_COL_BK_COST_TRANS: COL_COST_TRANSPORT_PRE_TAX,
        IN_COL_BK_VAT_TRANS: COL_VAT_TRANSPORT,
        IN_COL_BK_COST_SERV: COL_COST_SERVICE_PRE_TAX,
        IN_COL_BK_VAT_SERV: COL_VAT_SERVICE,
        IN_COL_BK_TOTAL: COL_TOTAL_AMOUNT,
    }
    bangKe = bangKe.rename(columns=rename_dict, errors='ignore')

    for col in [COL_COST_TRANSPORT_PRE_TAX, COL_COST_SERVICE_PRE_TAX, COL_VAT_TRANSPORT, COL_VAT_SERVICE]:
        if col in bangKe.columns: bangKe[col] = pd.to_numeric(bangKe[col], errors='coerce').fillna(0)

    if COL_COST_TRANSPORT_PRE_TAX in bangKe.columns:
        bangKe[COL_TOTAL_PRE_TAX] = bangKe[COL_COST_TRANSPORT_PRE_TAX] + bangKe.get(COL_COST_SERVICE_PRE_TAX, 0)
    if COL_VAT_TRANSPORT in bangKe.columns:
        bangKe[COL_TOTAL_VAT] = bangKe[COL_VAT_TRANSPORT] + bangKe.get(COL_VAT_SERVICE, 0)

    bangKe[COL_BOOKING_CODE_ORIG] = bangKe.get(COL_BOOKING_CODE, "")

    # Use Dynamic Output Order
    cols_to_keep = [OUT_COLS.get(k, '') for k in OUTPUT_ORDER_KEYS]
    cols_to_keep = [c for c in cols_to_keep if c] # Remove empty strings if any key missing
    
    bangKe = bangKe[[c for c in cols_to_keep if c in bangKe.columns]]

    # Map Report
    inv_group_map = bangKe.drop_duplicates(subset=[COL_INVOICE_NUM]).set_index(COL_INVOICE_NUM)[COL_GROUP_FUNCTION]
    report_df.insert(1, COL_GROUP_FUNCTION, report_df[COL_ADJUSTED_INVOICE_REPORT].map(inv_group_map.to_dict()))

    # --- Handle Discount Rows (Ensure Negative) ---
    if COL_SERVICE in bangKe.columns and COL_TOTAL_AMOUNT in bangKe.columns:
        is_discount = bangKe[COL_SERVICE].astype(str).str.contains('discount', case=False, na=False)
        bangKe.loc[is_discount, COL_TOTAL_AMOUNT] = -bangKe.loc[is_discount, COL_TOTAL_AMOUNT].abs()
        for col in [COL_COST_TRANSPORT_PRE_TAX, COL_VAT_TRANSPORT, COL_COST_SERVICE_PRE_TAX, COL_VAT_SERVICE, COL_TOTAL_PRE_TAX, COL_TOTAL_VAT]:
            if col in bangKe.columns:
                 bangKe.loc[is_discount, col] = -bangKe.loc[is_discount, col].abs()

    return bangKe, report_df


def generate_output_from_df(bangKe, report_df, file_bang_ke_original):
    """
    Nhận Dataframe đã xử lý (và chỉnh sửa), tạo các file Excel và Zip.
    """
    temp_dir = tempfile.mkdtemp()
    
    # Lưu file gốc tạm thời để dùng copy sheet Intro
    bk_path = os.path.join(temp_dir, "BangKe_Original.xlsx")
    with open(bk_path, "wb") as f: 
        file_bang_ke_original.seek(0)
        f.write(file_bang_ke_original.read())

    # --- Xuất File ---
    try:
        dt_series = pd.to_datetime(bangKe[COL_TIME], errors='coerce')
        month, year = int(dt_series.dt.month.max()), int(dt_series.dt.year.max())
        number_of_days = pd.Period(f'{year}-{month:02d}').days_in_month
    except:
        month, year, number_of_days = 0, 0, 0

    output_dir = os.path.join(temp_dir, "output")
    os.makedirs(output_dir, exist_ok=True)
    all_functions = bangKe[COL_GROUP_FUNCTION].dropna().unique()
    file_list_log = []

    # Files Con
    for func in all_functions:
        safe_func_name = str(func).strip().replace("/", "_").replace("\\", "_")
        filename = f"BK_GRAB_{safe_func_name}_{month}_{year}.xlsx"
        filepath = os.path.join(output_dir, filename)
        file_list_log.append(filename)
        
        df_func = bangKe[bangKe[COL_GROUP_FUNCTION] == func]
        df_ck = df_func[df_func[COL_PAYMENT_METHOD_INVOICE] == VAL_PAYMENT_METHOD_TRANSFER]
        df_tm = df_func[df_func[COL_PAYMENT_METHOD_INVOICE] != VAL_PAYMENT_METHOD_TRANSFER]
        
        writer = pd.ExcelWriter(filepath, engine='xlsxwriter')
        write_and_format_sheet_common(df_ck, f'1. BK {month}.{year} {func}'[:31], 'BẢNG KÊ CÁC CHUYẾN ĐI TRONG THÁNG (CHUYỂN KHOẢN)', writer, month, year, number_of_days, func)
        write_and_format_sheet_common(df_tm, '2.TM-CK', 'BẢNG KÊ CÁC CHUYẾN ĐI TRONG THÁNG (TM-CK)', writer, month, year, number_of_days, func)
        
        write_report_sheet(report_df[report_df[COL_ADJUSTED_INVOICE_REPORT].isin(df_ck[COL_INVOICE_NUM].unique())], 'DS Hoa don CK', 'BÁO CÁO HÓA ĐƠN (CK)', writer)
        write_report_sheet(report_df[report_df[COL_ADJUSTED_INVOICE_REPORT].isin(df_tm[COL_INVOICE_NUM].unique())], 'DS Hoa don TM-CK', 'BÁO CÁO HÓA ĐƠN (TM-CK)', writer)
        writer.close()

    # Master File
    master_filename = f"BK_GRAB_MASTER_{month}_{year}.xlsx"
    master_path = os.path.join(output_dir, master_filename)
    file_list_log.append(f"MASTER/{master_filename}") 
    writer_master = pd.ExcelWriter(master_path, engine='xlsxwriter')
    
    master_ck = bangKe[bangKe[COL_PAYMENT_METHOD_INVOICE] == VAL_PAYMENT_METHOD_TRANSFER]
    master_tm = bangKe[bangKe[COL_PAYMENT_METHOD_INVOICE] != VAL_PAYMENT_METHOD_TRANSFER]

    write_and_format_sheet_common(master_ck, '1. BK', 'BẢNG KÊ TỔNG HỢP', writer_master, month, year, number_of_days, "TẤT CẢ")
    write_and_format_sheet_common(master_tm, '2.TM-CK', 'BẢNG KÊ TỔNG HỢP (TM-CK)', writer_master, month, year, number_of_days, "TẤT CẢ")
    
    write_report_sheet(report_df[report_df[COL_ADJUSTED_INVOICE_REPORT].isin(master_ck[COL_INVOICE_NUM].unique())], 'DS Hoa don CK', 'BÁO CÁO HÓA ĐƠN TỔNG HỢP (CK)', writer_master)
    write_report_sheet(report_df[report_df[COL_ADJUSTED_INVOICE_REPORT].isin(master_tm[COL_INVOICE_NUM].unique())], 'DS Hoa don TM-CK', 'BÁO CÁO HÓA ĐƠN TỔNG HỢP (TM-CK)', writer_master)
    writer_master.close()
    
    copy_intro_sheet(bk_path, master_path)

    # Zip
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for root, dirs, files in os.walk(output_dir):
            for file in files:
                zip_file.write(os.path.join(root, file), file)
    
    shutil.rmtree(temp_dir)
    zip_buffer.seek(0)
    return zip_buffer, file_list_log

def distribute_pdfs_logic(df_invoice_list, source_dir, target_base_dir):
    logs = []
    if not os.path.exists(source_dir): return [f"❌ Thư mục nguồn không tồn tại: {source_dir}"]
    if not os.path.exists(target_base_dir): os.makedirs(target_base_dir)
    
    available_files = os.listdir(source_dir)
    count_success, count_fail = 0, 0
    progress_bar = st.progress(0)
    total = len(df_invoice_list)
    
    # Debug info for first few failures
    debug_failures = []
    
    def clean_float_str(s):
        if not s or str(s).lower() == 'nan': return ""
        s = str(s).strip()
        if s.endswith('.0'): return s[:-2]
        return s

    for idx, row in df_invoice_list.iterrows():
        try:
            grp = row.get(COL_GROUP_FUNCTION, "UNKNOWN")
            
            # Clean inputs
            inv_num = clean_float_str(row.get(COL_ADJUSTED_INVOICE_REPORT, ""))
            mau_so = clean_float_str(row.get('Mẫu số', ""))
            ky_hieu = str(row.get('Ký hiệu', "")).strip()
            if ky_hieu.lower() == 'nan': ky_hieu = ""
            
            if not inv_num: continue
            
            # Construct patterns to try (Strict to Loose)
            patterns = []
            if mau_so and ky_hieu:
                patterns.append(f"{mau_so}_{ky_hieu}_{inv_num}")
            
            # Also try just Invoice Number (most common fallback)
            patterns.append(f"_{inv_num}_") # Middle of string
            patterns.append(f"{inv_num}_") # Start or specific part
            
            safe_grp = str(grp).strip().replace(" ", "_").replace("/", "_")
            target_group_path = os.path.join(target_base_dir, safe_grp)
            
            if not os.path.exists(target_group_path): os.makedirs(target_group_path)
            
            matched_file = None
            
            # 1. Try strict match first
            for p in patterns:
                for f in available_files:
                    if p in f:
                        matched_file = f
                        break
                if matched_file: break
            
            # 2. Last resort: Check if file *contains* the invoice number as a standalone segment
            # e.g. "123.pdf" or "..._123_..."
            if not matched_file:
                 for f in available_files:
                     if inv_num in f:
                         # Validate it's not a partial match like "1234" matching "123"
                         # Simple check: surround with common delimiters
                         # But for now, just matching inv_num is better than nothing. 
                         # Check if it is a whole word or surrounded by _
                         if f"_{inv_num}_" in f or f.startswith(f"{inv_num}_"):
                             matched_file = f
                             break
            
            if matched_file:
                shutil.copy2(os.path.join(source_dir, matched_file), os.path.join(target_group_path, matched_file))
                count_success += 1
            else: 
                count_fail += 1
                if len(debug_failures) < 5:
                    debug_failures.append(f"Inv: {inv_num} | Pat: {patterns}")
        except: count_fail += 1
        if idx % 10 == 0: progress_bar.progress(min((idx + 1) / total, 1.0))
            
    progress_bar.progress(1.0)
    logs.append(f"✅ Hoàn tất! Copy thành công: {count_success}, Không tìm thấy: {count_fail}")
    if debug_failures:
        logs.append(f"🔍 Debug (5 lỗi đầu): {'; '.join(debug_failures)}")
    return logs

def distribute_all_files_logic(df_processed, df_report, source_pdf_dir, target_root_dir, file_bang_ke_original, file_function_mapping):
    """
    Hàm tổng hợp: Phân phối Excel, Email và PDF vào từng folder theo Group Function.
    """
    logs = []
    if not os.path.exists(source_pdf_dir): return [f"❌ Thư mục nguồn PDF không tồn tại: {source_pdf_dir}"]
    os.makedirs(target_root_dir, exist_ok=True)
    
    # 1. Prepare Intro Sheet Source
    temp_dir = tempfile.mkdtemp()
    bk_path_temp = os.path.join(temp_dir, "BangKe_Original.xlsx")
    try:
        file_bang_ke_original.seek(0)
        with open(bk_path_temp, "wb") as f: 
            f.write(file_bang_ke_original.read())
    except Exception as e:
        return [f"❌ Lỗi sao chép file Bảng Kê gốc: {e}"]

    # 2. Load Email Template & Mapping
    try:
        with open("templates/grab_invoice_email.html", "r", encoding="utf-8") as f:
            email_template = f.read()
    except:
        email_template = "<html><body><p>Template not found</p></body></html>"
        logs.append("⚠️ Không tìm thấy template email, dùng mặc định.")

    email_map = get_email_mapping_from_upload(file_function_mapping) if file_function_mapping else {}

    # 3. Get Groups
    if COL_GROUP_FUNCTION not in df_processed.columns:
        return ["❌ Không tìm thấy cột Group Function trong dữ liệu."]
    
    all_funcs = df_processed[COL_GROUP_FUNCTION].dropna().unique()
    available_pdfs = os.listdir(source_pdf_dir)
    
    progress_bar = st.progress(0)
    total_steps = len(all_funcs)
    
    count_excel, count_email, count_pdf = 0, 0, 0
    
    # Date info for Excel/Email
    try:
        dt_s = pd.to_datetime(df_processed[COL_TIME], errors='coerce')
        month, year = int(dt_s.dt.month.max()), int(dt_s.dt.year.max())
        number_of_days = pd.Period(f'{year}-{month:02d}').days_in_month
        month_year_str = f"{month}/{year}"
    except:
        month, year, number_of_days = 0, 0, 0
        month_year_str = "MM/YYYY"

    for idx, func in enumerate(all_funcs):
        safe_func_name = str(func).strip().replace("/", "_").replace("\\", "_")
        # Create Group Folder
        group_dir = os.path.join(target_root_dir, safe_func_name)
        os.makedirs(group_dir, exist_ok=True)
        
        # --- A. GENERATE EXCEL ---
        try:
            excel_filename = f"BK_GRAB_{safe_func_name}_{month}_{year}.xlsx"
            excel_path = os.path.join(group_dir, excel_filename)
            
            df_func = df_processed[df_processed[COL_GROUP_FUNCTION] == func]
            df_ck = df_func[df_func[COL_PAYMENT_METHOD_INVOICE] == VAL_PAYMENT_METHOD_TRANSFER]
            df_tm = df_func[df_func[COL_PAYMENT_METHOD_INVOICE] != VAL_PAYMENT_METHOD_TRANSFER]
            
            writer = pd.ExcelWriter(excel_path, engine='xlsxwriter')
            write_and_format_sheet_common(df_ck, f'1. BK {month}.{year} {func}'[:31], 'BẢNG KÊ CÁC CHUYẾN ĐI TRONG THÁNG (CHUYỂN KHOẢN)', writer, month, year, number_of_days, func)
            write_and_format_sheet_common(df_tm, '2.TM-CK', 'BẢNG KÊ CÁC CHUYẾN ĐI TRONG THÁNG (TM-CK)', writer, month, year, number_of_days, func)
            
            write_report_sheet(df_report[df_report[COL_ADJUSTED_INVOICE_REPORT].isin(df_ck[COL_INVOICE_NUM].unique())], 'DS Hoa don CK', 'BÁO CÁO HÓA ĐƠN (CK)', writer)
            write_report_sheet(df_report[df_report[COL_ADJUSTED_INVOICE_REPORT].isin(df_tm[COL_INVOICE_NUM].unique())], 'DS Hoa don TM-CK', 'BÁO CÁO HÓA ĐƠN (TM-CK)', writer)
            writer.close()
            
            # Copy Intro Sheet (Reload workbook not needed for copy_intro_sheet impl above, it uses openpyxl load)
            # But copy_intro_sheet takes paths.
            count_excel += 1
        except Exception as e:
            logs.append(f"❌ Lỗi tạo Excel cho {func}: {e}")

        # --- B. GENERATE EMAIL ---
        try:
            # Logic lookup email similar to Tab 6
            l_key = normalize_func_name(str(func))
            e_info = email_map.get(l_key, {})
            
            # Fallback to Group Name if needed (df_func usually has consistent Group per Function? Not always, but let's try)
            if not e_info.get('to'):
                 # Try finding via one of the groups in this function bundle
                 sample_grp = df_func[COL_GROUP].dropna().iloc[0] if not df_func[COL_GROUP].dropna().empty else ""
                 if sample_grp:
                     l_key_g = normalize_func_name(str(sample_grp))
                     e_info_g = email_map.get(l_key_g, {})
                     if e_info_g.get('to'): e_info = e_info_g
            
            e_to = e_info.get('to') or "user@example.com"
            e_cc = e_info.get('cc', "")
            r_name = extract_name_from_email(e_to)
            
            # Count Invoices
            inv_cnt = 0
            if COL_BOOKING_CODE in df_func.columns:
                inv_raw = df_func[COL_BOOKING_CODE].dropna().astype(str).str.strip()
                inv_cnt = inv_raw[inv_raw != ''].nunique()
            
            if inv_cnt > 0:
                d_map = {
                    'recipient_name': r_name,
                    'sender_name': 'Duyên',
                    'month_year': month_year_str,
                    'invoice_count': inv_cnt,
                    'vendor_code': 'VENDOR_CODE',
                    'po_number': 'PO_NUMBER',
                    'invoice_date': 'DD-MM-YYYY',
                    'cc_list': e_cc
                }
                
                final_html = email_template
                for k, v in d_map.items():
                    final_html = final_html.replace(f"{{{k}}}", str(v))
                
                safe_e = e_to.replace('/', '_').replace('\\', '_')
                # Filename format: email_{Group_Function}_{Email}.html
                email_filename = f"email_{safe_func_name}_{safe_e}.html"
                with open(os.path.join(group_dir, email_filename), "w", encoding="utf-8") as f:
                    f.write(final_html)
                count_email += 1
        except Exception as e:
            logs.append(f"⚠️ Lỗi tạo Email cho {func}: {e}")

        # --- C. COPY PDFS ---
        # Reuse logic from distribute_pdfs_logic but scoped to this group
        # Helper for cleaning
        def clean_float_str(s):
            if not s or str(s).lower() == 'nan': return ""
            s = str(s).strip()
            if s.endswith('.0'): return s[:-2]
            return s
        
        # Map Invoice -> Payment Method from Data (normalize keys)
        inv_pay_map = {}
        has_ck = False
        has_tm = False
        
        for idx_bk, row_bk in df_func.iterrows():
            raw_inv = clean_float_str(row_bk.get(COL_INVOICE_NUM, ""))
            pm = row_bk.get(COL_PAYMENT_METHOD_INVOICE, "")
            if raw_inv:
                inv_pay_map[raw_inv] = pm
            
            # Check existence
            if pm == VAL_PAYMENT_METHOD_TRANSFER: has_ck = True
            else: has_tm = True

        # Prepare subfolders (Only create if needed)
        dir_ck = os.path.join(group_dir, "HoaDon CK")
        dir_tm = os.path.join(group_dir, "HoaDon TM CK")
        
        if has_ck: os.makedirs(dir_ck, exist_ok=True)
        if has_tm: os.makedirs(dir_tm, exist_ok=True)

        # Iterate rows in this function only
        for _, row in df_report[df_report[COL_GROUP_FUNCTION] == func].iterrows():
            try:
                inv_num = clean_float_str(row.get(COL_ADJUSTED_INVOICE_REPORT, ""))
                mau_so = clean_float_str(row.get('Mẫu số', ""))
                ky_hieu = str(row.get('Ký hiệu', "")).strip()
                if ky_hieu.lower() == 'nan': ky_hieu = ""
                
                if not inv_num: continue
                
                patterns = []
                if mau_so and ky_hieu: patterns.append(f"{mau_so}_{ky_hieu}_{inv_num}")
                patterns.append(f"_{inv_num}_")
                patterns.append(f"{inv_num}_")
                
                matched = None
                for p in patterns:
                    for f in available_pdfs:
                        if p in f:
                            matched = f; break
                    if matched: break
                
                # 2. Relaxed matching
                if not matched:
                    for f in available_pdfs:
                        # Check if invoice number is in file name
                        if inv_num in f:
                            # Heuristic: if invoice number is long enough (>4 chars), assume it's unique enough
                            if len(inv_num) > 4:
                                matched = f; break
                            else:
                                # Short invoice number, be stricter
                                if f"_{inv_num}_" in f or f.startswith(f"{inv_num}_") or f.endswith(f"_{inv_num}.pdf"):
                                    matched = f; break
                
                if matched:
                    # Determine target subfolder based on Payment Method
                    pay_method = inv_pay_map.get(inv_num, "")
                    target_sub = None
                    
                    if pay_method == VAL_PAYMENT_METHOD_TRANSFER:
                        if has_ck: target_sub = dir_ck
                    else:
                        if has_tm: target_sub = dir_tm
                    
                    if target_sub:
                        shutil.copy2(os.path.join(source_pdf_dir, matched), os.path.join(target_sub, matched))
                        count_pdf += 1
            except: pass

        progress_bar.progress((idx + 1) / total_steps)

    # --- B. GENERATE MASTER EXCEL ---
    try:
        master_filename = f"BK_GRAB_MASTER_{month}_{year}.xlsx"
        master_path = os.path.join(target_root_dir, master_filename)
        
        # Filter data for Master
        master_ck = df_processed[df_processed[COL_PAYMENT_METHOD_INVOICE] == VAL_PAYMENT_METHOD_TRANSFER]
        master_tm = df_processed[df_processed[COL_PAYMENT_METHOD_INVOICE] != VAL_PAYMENT_METHOD_TRANSFER]

        writer_master = pd.ExcelWriter(master_path, engine='xlsxwriter')
        
        write_and_format_sheet_common(master_ck, '1. BK', 'BẢNG KÊ TỔNG HỢP', writer_master, month, year, number_of_days, "TẤT CẢ")
        write_and_format_sheet_common(master_tm, '2.TM-CK', 'BẢNG KÊ TỔNG HỢP (TM-CK)', writer_master, month, year, number_of_days, "TẤT CẢ")
        
        write_report_sheet(df_report[df_report[COL_ADJUSTED_INVOICE_REPORT].isin(master_ck[COL_INVOICE_NUM].unique())], 'DS Hoa don CK', 'BÁO CÁO HÓA ĐƠN TỔNG HỢP (CK)', writer_master)
        write_report_sheet(df_report[df_report[COL_ADJUSTED_INVOICE_REPORT].isin(master_tm[COL_INVOICE_NUM].unique())], 'DS Hoa don TM-CK', 'BÁO CÁO HÓA ĐƠN TỔNG HỢP (TM-CK)', writer_master)
        writer_master.close()
        
        # Copy Intro Sheet to Master
        copy_intro_sheet(bk_path_temp, master_path)
        logs.append(f"✅ Đã tạo Master File: {master_filename}")
        
    except Exception as e:
        logs.append(f"❌ Lỗi tạo Master File: {e}")

    shutil.rmtree(temp_dir)
    logs.append(f"✅ HOÀN TẤT TOÀN BỘ!")
    logs.append(f"📊 Thống kê: {count_excel} Excel, {count_email} Email, {count_pdf} PDF đã được phân phối.")
    return logs

# ==========================================
# 4. GIAO DIỆN STREAMLIT
# ==========================================

st.set_page_config(page_title="Grab Admin Tool Pro", layout="wide", initial_sidebar_state="expanded")
st.title("🚗 Grab Admin Tool")

# --- SIDEBAR ---
with st.sidebar:
    # st.image("https://cdn.haitrieu.com/wp-content/uploads/2022/10/Logo-Coca-Cola.png", width=100)
    st.title("🗂️ Menu Điều Khiển")
    
    with st.expander("ℹ️ Hướng dẫn nhanh", expanded=False):
        st.markdown("""
        1. **Tải dữ liệu**: Upload đủ 5 file yêu cầu bên dưới.
        2. **Xử lý**: Qua tab "Xử Lý Dữ Liệu" bấm nút chạy.
        3. **Kết quả**: Tải file Zip hoặc xem Dashboard phân tích.
        """)
    
    st.markdown("---")
    st.markdown("### 1️⃣ Dữ Liệu Grab (Input)")
    up_bang_ke = st.file_uploader("Bảng Kê Chi Tiết (.xlsx)", type=['xlsx'], help="File Excel chứa dữ liệu chuyến đi chi tiết")
    up_express = st.file_uploader("Dữ liệu Express (.csv)", type=['csv'], help="File CSV chi tiết dịch vụ giao hàng")
    up_transport = st.file_uploader("Dữ liệu Transport (.csv)", type=['csv'], help="File CSV chi tiết dịch vụ di chuyển")
    
    st.markdown("### 2️⃣ Dữ Liệu Cấu Hình")
    up_function = st.file_uploader("Phân Quyền Nhóm (.xlsx)", type=['xlsx'], help="File quy định nhóm phê duyệt")
    up_report = st.file_uploader("DS Hóa Đơn/Report (.xlsx)", type=['xlsx'], help="File theo dõi tình trạng hóa đơn")
    
    files_list = [up_bang_ke, up_express, up_transport, up_function, up_report]
    uploaded_count = sum([bool(f) for f in files_list])
    
    st.markdown("---")
    st.markdown("### 📊 Trạng Thái")
    if uploaded_count == 5:
        st.success("✅ Đã tải đủ dữ liệu!")
        st.markdown("**Sẵn sàng xử lý!**")
        if 'files_ok' not in st.session_state: st.session_state['files_ok'] = True
        files_ok = True
    else:
        st.progress(uploaded_count / 5)
        st.warning(f"⚠️ Đã tải {uploaded_count}/5 file. Vui lòng bổ sung.")
        if 'files_ok' not in st.session_state: st.session_state['files_ok'] = False
        files_ok = False
        
    st.caption(f"Phiên bản: 1.1.0 | {datetime.now().strftime('%d/%m/%Y')}")

# --- MAIN ---
if not files_ok:
    st.info("👋 Vui lòng tải lên đầy đủ dữ liệu để bắt đầu.")
else:
    tab_settings, tab_process, tab_filter, tab_pdf, tab_dashboard, tab_email, tab_draft = st.tabs([
        "⚙️ Cấu Hình",
        "⚙️ Xử Lý Dữ Liệu", 
        "🔎 Tra Cứu & Lọc", 
        "📂 Phân Phối & Đóng Gói",
        "📊 Dashboard",
        "✉️ Email Tool",
        "📧 Gửi Email (Draft)"
    ])
    
    # --- TAB 1: SETTINGS ---
    with tab_settings:
        st.header("⚙️ Cấu Hình Hệ Thống")
        st.markdown("Thay đổi các cài đặt và ánh xạ cột mà không cần chỉnh sửa code.")
        
        with st.expander("1. Cài Đặt Chung (General & Skiprows)", expanded=True):
            st.markdown("**Cấu hình dòng bắt đầu đọc (Skiprows):**")
            c1, c2 = st.columns(2)
            new_bk_sheet = c1.text_input("Tên Sheet Bảng Kê", value=BANG_KE_SHEET_NAME)
            new_skip_bk = c2.number_input("Bảng Kê (Skiprows)", value=SKIPROWS_BANG_KE, min_value=0)
            
            c3, c4 = st.columns(2)
            new_skip_express = c3.number_input("Dữ Liệu Express (Skiprows)", value=SKIPROWS_EXPRESS, min_value=0)
            new_skip_transport = c4.number_input("Dữ Liệu Transport (Skiprows)", value=SKIPROWS_TRANSPORT, min_value=0)
            
            # Raw Data Previews
            st.markdown("---")
            st.subheader("👀 Xem Trước Dữ Liệu Thô (Raw Previews)")
            
            p1, p2, p3 = st.tabs(["1. Bảng Kê (Excel)", "2. Express (CSV)", "3. Transport (CSV)"])
            
            with p1:
                if up_bang_ke:
                    try:
                        raw_df = pd.read_excel(up_bang_ke, sheet_name=new_bk_sheet, skiprows=new_skip_bk, nrows=5)
                        st.dataframe(raw_df)
                        st.caption(f"File: {up_bang_ke.name} | Sheet: {new_bk_sheet} | Skiprows: {new_skip_bk}")
                    except Exception as e:
                        st.error(f"Lỗi đọc file Bảng Kê: {e}")
                else:
                    st.info("Chưa tải file Bảng Kê.")
            
            with p2:
                if up_express:
                    try:
                        raw_express = pd.read_csv(up_express, skiprows=new_skip_express, nrows=5)
                        st.dataframe(raw_express)
                        st.caption(f"File: {up_express.name} | Skiprows: {new_skip_express}")
                    except Exception as e:
                        st.error(f"Lỗi đọc file Express: {e}")
                else:
                    st.info("Chưa tải file Express.")

            with p3:
                if up_transport:
                    try:
                        raw_transport = pd.read_csv(up_transport, skiprows=new_skip_transport, nrows=5)
                        st.dataframe(raw_transport)
                        st.caption(f"File: {up_transport.name} | Skiprows: {new_skip_transport}")
                    except Exception as e:
                        st.error(f"Lỗi đọc file Transport: {e}")
                else:
                    st.info("Chưa tải file Transport.")

        with st.expander("2. Ánh Xạ Cột Đầu Vào (Input Mappings)", expanded=False):
            st.markdown("Định nghĩa tên cột trong file Excel đầu vào tương ứng với biến hệ thống.")
            
            # Split mappings by category
            bk_cols = {k: v for k, v in IN_COLS.items() if 'IN_COL_BK_' in k}
            et_cols = {k: v for k, v in IN_COLS.items() if 'IN_COL_ET_' in k}
            func_cols = {k: v for k, v in IN_COLS.items() if 'IN_COL_FUNC_' in k}
            
            st.subheader("2.1. Bảng Kê (Excel)")
            df_in_bk = pd.DataFrame(list(bk_cols.items()), columns=['System Variable', 'Excel Column Header'])
            edited_in_bk = st.data_editor(df_in_bk, key="edit_bk", num_rows="dynamic", use_container_width=True)
            
            st.subheader("2.2. Express & Transport (CSV)")
            df_in_et = pd.DataFrame(list(et_cols.items()), columns=['System Variable', 'CSV Column Header'])
            edited_in_et = st.data_editor(df_in_et, key="edit_et", num_rows="dynamic", use_container_width=True)
            
            st.subheader("2.3. Function List (Excel)")
            df_in_func = pd.DataFrame(list(func_cols.items()), columns=['System Variable', 'Excel Column Header'])
            edited_in_func = st.data_editor(df_in_func, key="edit_func", num_rows="dynamic", use_container_width=True)

        with st.expander("3. Tên Cột Đầu Ra (Output Columns)", expanded=False):
            st.markdown("Định nghĩa tên cột sẽ hiển thị trong file báo cáo Excel.")
            df_out_cols = pd.DataFrame(list(OUT_COLS.items()), columns=['System Variable', 'Report Column Name'])
            edited_out_cols = st.data_editor(df_out_cols, num_rows="dynamic", use_container_width=True)

        with st.expander("4. Chọn Cột Xuất Báo Cáo (Output Selection)", expanded=False):
            st.markdown("Chọn và sắp xếp thứ tự các cột sẽ xuất hiện trong file cuối cùng.")
            all_out_keys = list(OUT_COLS.keys())
            current_selection = [k for k in OUTPUT_ORDER_KEYS if k in all_out_keys]
            
            selected_keys = st.multiselect(
                "Chọn cột (Thứ tự hiển thị dựa trên thứ tự bạn chọn/kéo thả trong danh sách này)",
                options=all_out_keys,
                default=current_selection
            )
            
            # Preview Output
            if selected_keys:
                dummy_cols = [edited_out_cols[edited_out_cols['System Variable'] == k]['Report Column Name'].values[0] 
                              if not edited_out_cols[edited_out_cols['System Variable'] == k].empty else k 
                              for k in selected_keys]
                st.caption("Thứ tự cột dự kiến:")
                st.dataframe(pd.DataFrame(columns=dummy_cols))
            
        if st.button("💾 Lưu Cấu Hình", type="primary"):
            new_config = CONFIG.copy() if CONFIG else {}
            if 'general' not in new_config: new_config['general'] = {}
            if 'input_columns' not in new_config: new_config['input_columns'] = {}
            if 'output_columns' not in new_config: new_config['output_columns'] = {}
            
            # Update General
            new_config['general']['BANG_KE_SHEET_NAME'] = new_bk_sheet
            new_config['general']['SKIPROWS_BANG_KE'] = int(new_skip_bk)
            new_config['general']['SKIPROWS_EXPRESS'] = int(new_skip_express)
            new_config['general']['SKIPROWS_TRANSPORT'] = int(new_skip_transport)
            
            # Update Inputs (Merge all tables back)
            new_in_cols = {}
            # Helper to convert df back to dict
            for df_res in [edited_in_bk, edited_in_et, edited_in_func]:
                 if not df_res.empty:
                    # Column 0 is Key, Column 1 is Value
                    for _, row in df_res.iterrows():
                         new_in_cols[row[0]] = row[1]
            
            new_config['input_columns'] = new_in_cols
            
            # Update Outputs
            new_out_cols = dict(zip(edited_out_cols['System Variable'], edited_out_cols['Report Column Name']))
            new_config['output_columns'] = new_out_cols
            
            # Update Order
            new_config['output_order'] = selected_keys
            
            save_config(new_config)
            st.success("✅ Đã lưu cấu hình! Vui lòng khởi động lại ứng dụng (Rerun) để áp dụng thay đổi.")

    # --- TAB 2: PROCESS ---
    with tab_process:
        st.header("Xử Lý Dữ Liệu & Tạo Báo Cáo")
        if st.button("🚀 1. Xử Lý & Xem Trước Dữ Liệu", type="primary"):
            with st.spinner("Đang đọc và xử lý dữ liệu..."):
                try:
                    up_bang_ke.seek(0); up_express.seek(0); up_transport.seek(0); up_function.seek(0); up_report.seek(0)
                    df_res, df_rep_res = process_input_data(up_bang_ke, up_express, up_transport, up_function, up_report)
                    st.session_state['df_preview'] = df_res
                    st.session_state['df_report_mapped'] = df_rep_res
                    st.success("✅ Đã xử lý xong! Vui lòng kiểm tra và chỉnh sửa bên dưới nếu cần.")
                    st.session_state['df_processed'] = df_res
                except Exception as e:
                    st.error(f"❌ Lỗi: {str(e)}")

        if 'df_preview' in st.session_state:
            st.markdown("### 📝 Xem Trước & Chỉnh Sửa")
            edited_df = st.data_editor(st.session_state['df_preview'], num_rows="dynamic", use_container_width=True, height=500)
            st.session_state['df_preview'] = edited_df
            st.session_state['df_processed'] = edited_df
            st.markdown("---")
            if st.button("💾 2. Xuất Báo Cáo & Tải Về"):
                with st.spinner("Đang tạo file Excel..."):
                    try:
                        up_bang_ke.seek(0)
                        zip_result, file_logs = generate_output_from_df(st.session_state['df_preview'], st.session_state['df_report_mapped'], up_bang_ke)
                        st.session_state['zip_result'] = zip_result
                        st.session_state['file_logs'] = file_logs
                        st.success("✅ Đã tạo file thành công!")
                    except Exception as e:
                        st.error(f"❌ Lỗi khi tạo file: {str(e)}")

        if 'zip_result' in st.session_state:
            st.markdown("---")
            c1, c2 = st.columns(2)
            with c1:
                st.download_button(label="Tải xuống trọn bộ (ZIP)", data=st.session_state['zip_result'], file_name=f"Grab_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.zip", mime="application/zip", type="primary")
            with c2:
                with st.container(height=200):
                    for f in st.session_state['file_logs']: st.code(f"📄 {f}", language="text")

    import plotly.express as px

    # --- TAB 3: FILTER ---
    with tab_filter:
        if 'df_processed' in st.session_state:
            df = st.session_state['df_processed'].copy()
            st.header("🔎 Tra Cứu & Lọc Nâng Cao")
            
            # --- 1. FILTER SECTION ---
            with st.expander("🛠️ BỘ LỌC DỮ LIỆU (ADVANCED FILTERS)", expanded=True):
                # Reset Button Logic
                if st.button("🔄 Reset Bộ Lọc", type="secondary"):
                    keys_to_reset = [
                        "filter_group", "filter_group_func", "filter_service", "filter_payment", "filter_city",
                        "filter_invoice", "filter_emp_name", "filter_location", "filter_booking",
                        "filter_general", "filter_date"
                    ]
                    for key in keys_to_reset:
                        if key in st.session_state:
                            del st.session_state[key]
                    st.rerun()

                # Row 1: Common Categorical Filters
                c1, c2, c3, c4 = st.columns(4)
                
                # Filter by Group (Multiselect)
                all_groups = sorted(list(df[COL_GROUP].dropna().unique())) if COL_GROUP in df.columns else []
                selected_groups = c1.multiselect("Nhóm:", all_groups, key="filter_group")

                # Filter by Group Function (Multiselect)
                all_group_funcs = sorted(list(df[COL_GROUP_FUNCTION].dropna().unique())) if COL_GROUP_FUNCTION in df.columns else []
                selected_group_funcs = c1.multiselect("Nhóm Chức Năng (Group Function):", all_group_funcs, key="filter_group_func")
                
                # Filter by Service (Multiselect)
                all_services = sorted(list(df[COL_SERVICE].dropna().unique())) if COL_SERVICE in df.columns else []
                selected_services = c2.multiselect("Loại Dịch Vụ:", all_services, key="filter_service")
                
                # Filter by Payment Type (Multiselect)
                pay_col = COL_PAYMENT_METHOD_INVOICE if COL_PAYMENT_METHOD_INVOICE in df.columns else COL_PAYMENT_TYPE
                all_payments = sorted(list(df[pay_col].astype(str).dropna().unique())) if pay_col in df.columns else []
                selected_payments = c3.multiselect("Hình Thức Thanh Toán:", all_payments, key="filter_payment")

                # Filter by City (Multiselect)
                all_cities = sorted(list(df[COL_CITY].astype(str).dropna().unique())) if COL_CITY in df.columns else []
                selected_cities = c4.multiselect("Tỉnh/Thành phố:", all_cities, key="filter_city")

                # --- Dynamic Employee Filter ---
                # Calculate available employees based on Group/Service/City/GroupFunction selection
                temp_df = df.copy()
                if selected_groups:
                    temp_df = temp_df[temp_df[COL_GROUP].isin(selected_groups)]
                if selected_group_funcs:
                    temp_df = temp_df[temp_df[COL_GROUP_FUNCTION].isin(selected_group_funcs)]
                if selected_services:
                    temp_df = temp_df[temp_df[COL_SERVICE].isin(selected_services)]
                if selected_cities:
                    temp_df = temp_df[temp_df[COL_CITY].isin(selected_cities)]
                
                available_employees = sorted(list(temp_df[COL_EMPLOYEE_NAME].dropna().unique())) if COL_EMPLOYEE_NAME in temp_df.columns else []

                # Row 2: Search Fields
                c5, c6, c7, c8 = st.columns(4)
                
                search_invoice = c5.text_input("Mã Hóa Đơn:", key="filter_invoice")
                
                # Replaced Emp ID text input with Dynamic Multiselect for Employee Name
                selected_employees = c6.multiselect("Tên Nhân Viên (theo lọc):", available_employees, key="filter_emp_name")
                
                search_location = c7.text_input("Địa điểm (Pickup/Dropoff):", key="filter_location")
                search_booking = c8.text_input("Mã Chuyến Xe:", key="filter_booking")

                # Row 3: Date Range & General Search
                c9, c10 = st.columns([1, 3])
                
                # Ensure Date Object
                if 'Date_Obj' not in df.columns:
                    df['Date_Obj'] = pd.to_datetime(df[COL_TIME], dayfirst=True, errors='coerce')
                
                # Logic: Default to full range if no specific selection
                min_d = df['Date_Obj'].min().date() if not df['Date_Obj'].isnull().all() else datetime.now().date()
                max_d = df['Date_Obj'].max().date() if not df['Date_Obj'].isnull().all() else datetime.now().date()
                
                # Use session state to persist or reset date
                date_range = c9.date_input("Khoảng Thời Gian:", value=(min_d, max_d), key="filter_date")
                
                search_general = c10.text_input("Tìm kiếm chung (Mã NV, Mô tả...):", key="filter_general")

            # --- 2. APPLY FILTERS ---
            df_filtered = df.copy()
            
            # Categorical (List check)
            if selected_groups:
                df_filtered = df_filtered[df_filtered[COL_GROUP].isin(selected_groups)]
            if selected_group_funcs:
                df_filtered = df_filtered[df_filtered[COL_GROUP_FUNCTION].isin(selected_group_funcs)]
            if selected_services:
                df_filtered = df_filtered[df_filtered[COL_SERVICE].isin(selected_services)]
            if selected_payments:
                df_filtered = df_filtered[df_filtered[pay_col].isin(selected_payments)]
            if selected_cities:
                df_filtered = df_filtered[df_filtered[COL_CITY].isin(selected_cities)]
            
            # Employee Name Filter
            if selected_employees:
                df_filtered = df_filtered[df_filtered[COL_EMPLOYEE_NAME].isin(selected_employees)]

            # Text Search
            if search_invoice and COL_INVOICE_NUM in df.columns:
                df_filtered = df_filtered[df_filtered[COL_INVOICE_NUM].astype(str).str.contains(search_invoice, case=False, na=False)]
            
            if search_booking and COL_BOOKING_CODE in df.columns:
                df_filtered = df_filtered[df_filtered[COL_BOOKING_CODE].astype(str).str.contains(search_booking, case=False, na=False)]
            
            if search_location:
                loc_mask = pd.Series(False, index=df_filtered.index)
                if COL_PICKUP in df.columns:
                    loc_mask |= df_filtered[COL_PICKUP].astype(str).str.contains(search_location, case=False, na=False)
                if COL_DROPOFF in df.columns:
                    loc_mask |= df_filtered[COL_DROPOFF].astype(str).str.contains(search_location, case=False, na=False)
                df_filtered = df_filtered[loc_mask]

            if search_general:
                gen_mask = pd.Series(False, index=df_filtered.index)
                if COL_EMPLOYEE_ID in df.columns:
                    gen_mask |= df_filtered[COL_EMPLOYEE_ID].astype(str).str.contains(search_general, case=False, na=False)
                if COL_TRIP_PURPOSE in df.columns:
                    gen_mask |= df_filtered[COL_TRIP_PURPOSE].astype(str).str.contains(search_general, case=False, na=False)
                df_filtered = df_filtered[gen_mask]

            # Date Range (Apply ONLY if range is valid)
            if isinstance(date_range, tuple) and len(date_range) == 2:
                start_d, end_d = date_range
                
                # Check if filter matches full range (approximate check to allow "All Time" behavior)
                is_full_range = False
                if min_d and max_d:
                    if start_d <= min_d and end_d >= max_d:
                        is_full_range = True
                
                if not is_full_range:
                    # Only filter if NOT full range, OR explicit filter requested
                    # Filter: Keep rows within range. Rows with NaT are excluded by default in comparison.
                    # If we want to keep NaT when range is full, we skip this block or handle NaT.
                    
                    df_filtered = df_filtered[
                        (df_filtered['Date_Obj'].dt.date >= start_d) & 
                        (df_filtered['Date_Obj'].dt.date <= end_d)
                    ]
                # Else: If full range selected, we do NOTHING to df_filtered regarding date, 
                # so rows with NaT (invalid dates) are preserved.

            # --- 3. RESULTS TABLE ---
            st.markdown("---")
            c_res1, c_res2 = st.columns([3, 1])
            c_res1.markdown(f"### 📋 Kết quả tìm kiếm: {len(df_filtered):,} dòng")
            
            if not df_filtered.empty:
                # Download Button
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                    df_filtered.to_excel(writer, index=False, sheet_name='Filtered_Data')
                
                c_res2.download_button(
                    label="⬇️ Tải Excel",
                    data=output.getvalue(),
                    file_name="Grab_Filtered_Data.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="btn_download_filter"
                )
                
                st.dataframe(df_filtered, use_container_width=True, height=400)
                
                # --- 4. MINI DASHBOARD (ANALYSIS) ---
                st.markdown("---")
                st.subheader("📊 Dashboard Phân Tích Kết Quả Lọc")
                
                # Metrics
                f_amount = pd.to_numeric(df_filtered[COL_TOTAL_AMOUNT], errors='coerce').sum()
                f_vat = pd.to_numeric(df_filtered[COL_TOTAL_VAT], errors='coerce').sum()
                f_trips = len(df_filtered)
                
                m1, m2, m3 = st.columns(3)
                m1.metric("Tổng Chi Phí (Selection)", f"{f_amount:,.0f} đ")
                m2.metric("Tổng Thuế VAT", f"{f_vat:,.0f} đ")
                m3.metric("Số Lượng Chuyến", f"{f_trips:,}")
                
                # Charts
                fc1, fc2 = st.columns(2)
                
                with fc1:
                    # Pie Chart: Service Type
                    if COL_SERVICE in df_filtered.columns:
                        f_service = df_filtered.groupby(COL_SERVICE)[COL_TOTAL_AMOUNT].sum().reset_index()
                        fig_f_service = px.pie(f_service, values=COL_TOTAL_AMOUNT, names=COL_SERVICE, title="Tỷ Trọng Dịch Vụ (Filtered)", hole=0.4)
                        st.plotly_chart(fig_f_service, use_container_width=True)
                
                with fc2:
                    # Bar Chart: Group (if multiple groups selected or All)
                    if COL_GROUP in df_filtered.columns:
                        f_group = df_filtered.groupby(COL_GROUP)[COL_TOTAL_AMOUNT].sum().reset_index().sort_values(COL_TOTAL_AMOUNT, ascending=True).tail(10)
                        fig_f_group = px.bar(f_group, x=COL_TOTAL_AMOUNT, y=COL_GROUP, orientation='h', title="Top Nhóm Chi Tiêu (Filtered)", text_auto='.2s')
                        st.plotly_chart(fig_f_group, use_container_width=True)
                
                # Daily Trend Line
                st.markdown("**Xu Hướng Theo Ngày (Filtered)**")
                if 'Date_Only' in df_filtered.columns:
                    f_daily = df_filtered.groupby('Date_Only')[COL_TOTAL_AMOUNT].sum().reset_index()
                    fig_f_trend = px.line(f_daily, x='Date_Only', y=COL_TOTAL_AMOUNT, markers=True)
                    st.plotly_chart(fig_f_trend, use_container_width=True)

            else:
                st.warning("Không tìm thấy dữ liệu phù hợp với bộ lọc hiện tại.")

        else:
            st.info("Vui lòng chạy xử lý ở Tab '⚙️ Xử Lý Dữ Liệu' trước.")

    # --- TAB 4: PDF ---
    with tab_pdf:
        st.header("📂 Phân Phối & Đóng Gói Hồ Sơ (Deployment Ready)")
        st.markdown("""
        Chức năng này sẽ tự động thực hiện quy trình khép kín:
        1.  **Phân loại Excel**: Tạo file Excel Bảng Kê riêng cho từng Group Function.
        2.  **Tạo Master File**: Tổng hợp toàn bộ dữ liệu vào 1 file Master.
        3.  **Tạo Email Draft**: Tạo file `.eml` (Outlook) đính kèm sẵn Excel & PDF.
        4.  **Phân loại PDF**: Tìm và copy PDF hóa đơn vào thư mục (và file Zip đính kèm Email).
        5.  **Đóng gói**: Nén toàn bộ kết quả thành 1 file Zip để tải về.
        """)
        
        st.markdown("---")
        st.subheader("1. Dữ liệu PDF Đầu Vào")
        uploaded_pdf_zip = st.file_uploader("Tải lên file Zip chứa toàn bộ PDF Hóa Đơn", type=['zip'], help="Nén toàn bộ file PDF hóa đơn vào 1 file .zip và tải lên đây.")

        if 'df_processed' in st.session_state and uploaded_pdf_zip:
            if st.button("🚀 Bắt đầu Phân Phối & Đóng Gói", type="primary"):
                with st.spinner("Đang xử lý... Vui lòng đợi..."):
                    try:
                        # Create temporary directories
                        with tempfile.TemporaryDirectory() as temp_root:
                            temp_pdf_in = os.path.join(temp_root, "PDF_INPUT")
                            temp_out = os.path.join(temp_root, "OUTPUT")
                            os.makedirs(temp_pdf_in, exist_ok=True)
                            os.makedirs(temp_out, exist_ok=True)
                            
                            # 1. Extract PDF Zip
                            st.info("📦 Đang giải nén PDF...")
                            with zipfile.ZipFile(uploaded_pdf_zip, 'r') as zip_ref:
                                zip_ref.extractall(temp_pdf_in)
                            
                            # Flatten: Move all PDFs from subfolders to root of temp_pdf_in
                            # This fixes issue where zip contains a folder (e.g. "MyPDFs/file.pdf")
                            pdf_moved_count = 0
                            for root, dirs, files in os.walk(temp_pdf_in):
                                if root == temp_pdf_in: continue # Skip root
                                for file in files:
                                    if file.lower().endswith('.pdf'):
                                        src_p = os.path.join(root, file)
                                        dst_p = os.path.join(temp_pdf_in, file)
                                        # Avoid overwrite if name collision (simple rename)
                                        if os.path.exists(dst_p):
                                            base, ext = os.path.splitext(file)
                                            dst_p = os.path.join(temp_pdf_in, f"{base}_{uuid.uuid4().hex[:4]}{ext}")
                                        shutil.move(src_p, dst_p)
                                        pdf_moved_count += 1
                            
                            # Debug: Count PDFs in root
                            pdfs_in_root = [f for f in os.listdir(temp_pdf_in) if f.lower().endswith('.pdf')]
                            st.write(f"ℹ️ Tìm thấy {len(pdfs_in_root)} file PDF trong file nén.")
                            if len(pdfs_in_root) == 0:
                                st.warning("⚠️ Không tìm thấy file PDF nào trong file nén. Vui lòng kiểm tra lại cấu trúc file Zip.")

                            # 2. Run Distribution Logic
                            st.info("⚙️ Đang phân phối dữ liệu & Tạo Email Draft...")
                            
                            # Reuse distribute_all_files_logic
                            logs = distribute_all_files_logic(
                                st.session_state['df_processed'],
                                st.session_state.get('df_report_mapped', pd.DataFrame()), 
                                temp_pdf_in,
                                temp_out,
                                up_bang_ke,
                                up_function
                            )
                            
                            # 2.5 Run Create Email Draft for ALL Groups
                            # Since distribute_all_files_logic creates folders, we iterate temp_out to create drafts.
                            st.info("✉️ Đang tạo file Email Draft (.eml)...")
                            draft_logs = create_eml_draft(temp_out)
                            logs.extend(draft_logs)
                            
                            # 3. Zip Output
                            st.info("🗜️ Đang nén kết quả...")
                            
                            final_zip_buffer = io.BytesIO()
                            with zipfile.ZipFile(final_zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_final:
                                for root, dirs, files in os.walk(temp_out):
                                    for file in files:
                                        abs_path = os.path.join(root, file)
                                        rel_path = os.path.relpath(abs_path, temp_out)
                                        zip_final.write(abs_path, arcname=rel_path)
                            
                            final_zip_buffer.seek(0)
                            
                            st.success("✅ Xử lý hoàn tất!")
                            
                            # Show Logs
                            with st.expander("Chi tiết xử lý"):
                                for log in logs:
                                    if "❌" in log: st.error(log)
                                    elif "⚠️" in log: st.warning(log)
                                    else: st.text(log)

                            # Download Button
                            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                            st.download_button(
                                label="⬇️ Tải Xuống Kết Quả (Zip)",
                                data=final_zip_buffer,
                                file_name=f"KETQUA_PHAN_PHOI_{timestamp}.zip",
                                mime="application/zip"
                            )

                    except Exception as e:
                        st.error(f"❌ Có lỗi xảy ra: {e}")
                        st.exception(e)

        elif 'df_processed' not in st.session_state:
            st.warning("⚠️ Vui lòng chạy 'Xử Lý Dữ Liệu' (Tab 2) trước.")
        elif not uploaded_pdf_zip:
            st.info("ℹ️ Vui lòng tải lên file Zip chứa PDF để tiếp tục.")

    # --- TAB 5: DASHBOARD ---
    with tab_dashboard:
        if 'df_processed' in st.session_state:
            import plotly.graph_objects as go
            import plotly.express as px
            import pydeck as pdk

            df = st.session_state['df_processed'].copy()
            
            # Ensure numeric columns
            money_cols = [COL_TOTAL_AMOUNT, COL_TOTAL_VAT, COL_TOTAL_PRE_TAX, COL_COST_TRANSPORT_PRE_TAX, COL_VAT_TRANSPORT, COL_COST_SERVICE_PRE_TAX, COL_VAT_SERVICE, COL_SURCHARGE]
            for col in money_cols:
                if col in df.columns: df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
            
            # Date processing
            df['Date_Obj'] = pd.to_datetime(df[COL_TIME], dayfirst=True, errors='coerce')
            df['Date_Only'] = df['Date_Obj'].dt.date
            df['Month'] = df['Date_Obj'].dt.strftime('%m-%Y')
            df['Hour'] = df['Date_Obj'].dt.hour
            df['DayOfWeek'] = df['Date_Obj'].dt.day_name()
            
            # Sort DayOfWeek
            days_order = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
            df['DayOfWeek'] = pd.Categorical(df['DayOfWeek'], categories=days_order, ordered=True)

            st.header("📊 Dashboard Phân Tích Chi Phí Toàn Diện")
            st.markdown("---")
            
            # 1. Key Metrics (Enhanced Power BI Style)
            st.subheader("1. Tổng Quan & Phân Tích Chuyên Sâu (Power BI View)")
            
            # --- Row 1: KPI Cards ---
            k1, k2, k3, k4, k5 = st.columns(5)
            gross_spend = df[COL_TOTAL_AMOUNT].sum()
            total_vat = df[COL_TOTAL_VAT].sum()
            total_pre_tax = df[COL_TOTAL_PRE_TAX].sum()
            total_trips = len(df)
            avg_trip = gross_spend / total_trips if total_trips > 0 else 0
            
            # Calculate MoM (Month over Month) trend if old data available (Mockup logic or based on date range split)
            # For now, we show static metrics but formatted nicely
            k1.metric("💰 Tổng Chi (Gross)", f"{gross_spend:,.0f}", help="Tổng chi phí thực tế phải thanh toán")
            k2.metric("💵 Trước Thuế (Net)", f"{total_pre_tax:,.0f}", help="Chi phí chưa bao gồm VAT")
            k3.metric("💸 Thuế (VAT)", f"{total_vat:,.0f}", delta=f"{total_vat/gross_spend*100:.1f}% tỷ lệ", help="Tổng tiền thuế VAT")
            k4.metric("🚖 Số Chuyến", f"{total_trips:,.0f}", help="Tổng số chuyến xe phát sinh")
            k5.metric("🏷️ Trung Bình/Chuyến", f"{avg_trip:,.0f}", help="Chi phí trung bình cho mỗi chuyến đi")

            # --- Row 2: Deep Dive Visuals ---
            st.markdown("##### 🧩 Cấu Trúc Chi Phí & Phân Tích Hành Vi")
            v1, v2 = st.columns([2, 1])
            
            with v1:
                # Treemap: Group -> Service -> Amount
                if COL_GROUP in df.columns and COL_SERVICE in df.columns:
                     # Handle NaNs for Treemap (Critical to avoid None entries error)
                     df_treemap = df.copy()
                     df_treemap[COL_GROUP] = df_treemap[COL_GROUP].fillna("Unknown Group")
                     df_treemap[COL_SERVICE] = df_treemap[COL_SERVICE].fillna("Unknown Service")
                     
                     # Ensure positive values for treemap size (Plotly Treemap doesn't like negative values for size)
                     # We can use absolute value for size, but color by actual value
                     df_treemap['Abs_Amount'] = df_treemap[COL_TOTAL_AMOUNT].abs()
                     
                     fig_tree = px.treemap(
                         df_treemap, 
                         path=[px.Constant("Toàn Bộ"), COL_GROUP, COL_SERVICE], 
                         values='Abs_Amount', # Use absolute for size
                         color=COL_TOTAL_AMOUNT, # Use actual for color (Red for negative/discount)
                         color_continuous_scale='RdBu_r',
                         title="Bản Đồ Cấu Trúc Chi Phí (Treemap: Nhóm > Dịch Vụ)"
                     )
                     fig_tree.update_traces(textinfo="label+value+percent entry")
                     fig_tree.update_layout(margin=dict(t=30, l=0, r=0, b=0))
                     st.plotly_chart(fig_tree, use_container_width=True)
            
            with v2:
                # Scatter: Spend vs Trips (Employee Level)
                if COL_EMPLOYEE_NAME in df.columns:
                    emp_stats = df.groupby([COL_EMPLOYEE_NAME, COL_GROUP]).agg({
                        COL_TOTAL_AMOUNT: 'sum',
                        COL_BOOKING_CODE: 'count'
                    }).reset_index()
                    emp_stats['Avg_Cost'] = emp_stats[COL_TOTAL_AMOUNT] / emp_stats[COL_BOOKING_CODE]
                    
                    fig_scatter = px.scatter(
                        emp_stats, 
                        x=COL_BOOKING_CODE, 
                        y=COL_TOTAL_AMOUNT,
                        size='Avg_Cost', 
                        color=COL_GROUP,
                        hover_name=COL_EMPLOYEE_NAME,
                        title="Phân Tích Hành Vi Nhân Viên (Scatter)",
                        labels={COL_BOOKING_CODE: "Số Chuyến", COL_TOTAL_AMOUNT: "Tổng Chi"},
                        size_max=40
                    )
                    fig_scatter.update_layout(margin=dict(t=30, l=0, r=0, b=0), showlegend=False)
                    st.plotly_chart(fig_scatter, use_container_width=True)
            
            st.markdown("---")

            # 2. Daily & Hourly Trends (Separated Rows)
            st.markdown("##### 📅 Xu Hướng Theo Ngày (Chi Phí & Số Chuyến)")
            if not df.empty:
                daily_agg = df.groupby('Date_Only').agg({
                    COL_TOTAL_AMOUNT: 'sum',
                    COL_BOOKING_CODE: 'count'
                }).reset_index()
                
                fig_daily = go.Figure()
                # Bar: Cost
                fig_daily.add_trace(go.Bar(
                    x=daily_agg['Date_Only'], 
                    y=daily_agg[COL_TOTAL_AMOUNT],
                    name='Chi Phí',
                    text=daily_agg[COL_TOTAL_AMOUNT],
                    texttemplate='%{text:.2s}',
                    textposition='auto',
                    marker=dict(color=daily_agg[COL_TOTAL_AMOUNT], colorscale='Teal')
                ))
                # Line: Trips
                fig_daily.add_trace(go.Scatter(
                    x=daily_agg['Date_Only'],
                    y=daily_agg[COL_BOOKING_CODE],
                    name='Số Chuyến',
                    yaxis='y2',
                    mode='lines+markers',
                    line=dict(shape='spline', width=3, color='#FF6B6B'),
                    marker=dict(size=6)
                ))
                
                fig_daily.update_layout(
                    xaxis=dict(title='Ngày'),
                    yaxis=dict(title='Chi Phí (VND)', showgrid=False),
                    yaxis2=dict(title='Số Chuyến', overlaying='y', side='right', showgrid=False),
                    legend=dict(x=0, y=1.2, orientation='h'),
                    margin=dict(l=0, r=0, t=30, b=0),
                    height=400
                )
                st.plotly_chart(fig_daily, use_container_width=True)

            st.markdown("##### ⏰ Phân Bố Theo Khung Giờ (Chi Phí & Số Chuyến)")
            if 'Hour' in df.columns:
                hourly_agg = df.groupby('Hour').agg({
                    COL_TOTAL_AMOUNT: 'sum',
                    COL_BOOKING_CODE: 'count'
                }).reset_index()
                
                fig_hour = go.Figure()
                # Bar: Cost
                fig_hour.add_trace(go.Bar(
                    x=hourly_agg['Hour'],
                    y=hourly_agg[COL_TOTAL_AMOUNT],
                    name='Chi Phí',
                    text=hourly_agg[COL_TOTAL_AMOUNT],
                    texttemplate='%{text:.2s}',
                    textposition='auto',
                    marker=dict(color=hourly_agg[COL_TOTAL_AMOUNT], colorscale='Viridis')
                ))
                # Line: Trips
                fig_hour.add_trace(go.Scatter(
                    x=hourly_agg['Hour'],
                    y=hourly_agg[COL_BOOKING_CODE],
                    name='Số Chuyến',
                    yaxis='y2',
                    mode='lines+markers',
                    line=dict(shape='spline', width=3, color='#FF6B6B'),
                    marker=dict(size=6)
                ))
                
                fig_hour.update_layout(
                    xaxis=dict(title='Giờ trong ngày', dtick=1),
                    yaxis=dict(title='Chi Phí (VND)', showgrid=False),
                    yaxis2=dict(title='Số Chuyến', overlaying='y', side='right', showgrid=False),
                    legend=dict(x=0, y=1.2, orientation='h'),
                    margin=dict(l=0, r=0, t=30, b=0),
                    height=400
                )
                st.plotly_chart(fig_hour, use_container_width=True)

            # 3. Department & Service
            c1, c2 = st.columns(2)
            with c1:
                st.markdown("##### 🏢 Chi Phí Theo Phòng Ban")
                if COL_GROUP in df.columns:
                    group_cost = df.groupby(COL_GROUP)[COL_TOTAL_AMOUNT].sum().reset_index().sort_values(COL_TOTAL_AMOUNT, ascending=False).head(10)
                    fig_group = px.bar(
                        group_cost, 
                        x=COL_TOTAL_AMOUNT, 
                        y=COL_GROUP, 
                        orientation='h', 
                        text_auto='.2s', 
                        title="",
                        color=COL_TOTAL_AMOUNT,
                        color_continuous_scale='Blues'
                    )
                    st.plotly_chart(fig_group, use_container_width=True)
            with c2:
                st.markdown("##### 🚕 Tỷ Trọng Dịch Vụ")
                if COL_SERVICE in df.columns:
                    service_cost = df.groupby(COL_SERVICE)[COL_TOTAL_AMOUNT].sum().reset_index()
                    fig_service = px.pie(
                        service_cost, 
                        values=COL_TOTAL_AMOUNT, 
                        names=COL_SERVICE, 
                        title="", 
                        hole=0.4,
                        color_discrete_sequence=px.colors.sequential.RdBu
                    )
                    st.plotly_chart(fig_service, use_container_width=True)

            # 4. Advanced Heatmap
            st.markdown("---")
            st.subheader("🔥 Bản Đồ Nhiệt: Tần Suất Đặt Xe (Thứ vs Giờ)")
            if 'DayOfWeek' in df.columns and 'Hour' in df.columns:
                heatmap_data = df.pivot_table(index='DayOfWeek', columns='Hour', values=COL_BOOKING_CODE, aggfunc='count')
                
                # Ensure all hours 0-23 are present
                heatmap_data = heatmap_data.reindex(columns=range(24), fill_value=0)

                fig_heat = px.imshow(
                    heatmap_data, 
                    labels=dict(x="Giờ trong ngày", y="Thứ trong tuần", color="Số Chuyến"),
                    color_continuous_scale="Viridis",
                    aspect="auto",
                    text_auto=True
                )
                fig_heat.update_traces(xgap=2, ygap=2)
                fig_heat.update_xaxes(dtick=1, range=[-0.5, 23.5]) # Force 0-23 range
                st.plotly_chart(fig_heat, use_container_width=True)

            # 5. Employee Dashboard
            st.markdown("---")
            st.subheader("👤 Dashboard Chi Tiết Nhân Viên")
            
            all_emps = sorted(df[COL_EMPLOYEE_NAME].dropna().unique()) if COL_EMPLOYEE_NAME in df.columns else []
            selected_emp = st.selectbox("Chọn Nhân Viên để phân tích:", all_emps)
            
            if selected_emp:
                emp_df = df[df[COL_EMPLOYEE_NAME] == selected_emp]
                
                # Emp Metrics
                ec1, ec2, ec3 = st.columns(3)
                e_gross = emp_df[COL_TOTAL_AMOUNT].sum()
                e_trips = len(emp_df)
                e_avg = e_gross / e_trips if e_trips > 0 else 0
                
                ec1.metric("Tổng Chi Tiêu", f"{e_gross:,.0f} đ")
                ec2.metric("Tổng Số Chuyến", f"{e_trips}")
                ec3.metric("Trung Bình/Chuyến", f"{e_avg:,.0f} đ")
                
                # Emp Charts
                ec4, ec5 = st.columns(2)
                with ec4:
                    e_daily = emp_df.groupby('Date_Only')[COL_TOTAL_AMOUNT].sum().reset_index()
                    fig_e_daily = px.bar(e_daily, x='Date_Only', y=COL_TOTAL_AMOUNT, title="Chi Tiêu Theo Ngày")
                    st.plotly_chart(fig_e_daily, use_container_width=True)
                with ec5:
                    if COL_SERVICE in emp_df.columns:
                        e_service = emp_df.groupby(COL_SERVICE)[COL_TOTAL_AMOUNT].sum().reset_index()
                        fig_e_service = px.pie(e_service, values=COL_TOTAL_AMOUNT, names=COL_SERVICE, title="Dịch Vụ Sử Dụng", hole=0.4)
                        st.plotly_chart(fig_e_service, use_container_width=True)

            # 6. 3D Map & Location Pivot
            st.markdown("---")
            st.subheader("🗺️ Bản Đồ Phân Bố & Thống Kê Địa Điểm")
            
            # Map Controls
            if st.button("🔄 Reset Map View"):
                if 'map_view_state' in st.session_state:
                     del st.session_state['map_view_state']
                st.rerun()

            cm1, cm2 = st.columns([2, 1])
            
            # Prepare Map Data
            CITY_COORDS = {
                'Ho Chi Minh': {'lat': 10.8231, 'lon': 106.6297},
                'Hanoi': {'lat': 21.0285, 'lon': 105.8542},
                'Da Nang': {'lat': 16.0544, 'lon': 108.2022},
                'Can Tho': {'lat': 10.0452, 'lon': 105.7469},
                'Hai Phong': {'lat': 20.8449, 'lon': 106.6881},
                'Khanh Hoa': {'lat': 12.2388, 'lon': 109.1967},
                'Lam Dong': {'lat': 11.9404, 'lon': 108.4583},
                'Ba Ria - Vung Tau': {'lat': 10.3460, 'lon': 107.0843},
                'Quang Ninh': {'lat': 20.9599, 'lon': 107.0425},
                'Dong Nai': {'lat': 10.9422, 'lon': 106.8180},
                'Binh Duong': {'lat': 11.1601, 'lon': 106.6669},
                'Unknown': {'lat': None, 'lon': None}
            }
            
            if COL_CITY in df.columns:
                map_df = df.groupby(COL_CITY).agg({COL_TOTAL_AMOUNT: 'sum', COL_BOOKING_CODE: 'count'}).reset_index()
                # Ensure CITY_COORDS keys match map_df values (handle case/strip)
                map_df['City_Normalized'] = map_df[COL_CITY].astype(str).str.strip() # Normalize if needed
                
                # Simple fallback if key not found? map returns NaN
                map_df['lat'] = map_df['City_Normalized'].map(lambda x: CITY_COORDS.get(x, {}).get('lat'))
                map_df['lon'] = map_df['City_Normalized'].map(lambda x: CITY_COORDS.get(x, {}).get('lon'))
                
                # Debug if empty
                if map_df['lat'].isnull().all() and not map_df.empty:
                     st.warning(f"Không tìm thấy tọa độ cho các thành phố: {list(map_df[COL_CITY].unique())}. Vui lòng kiểm tra lại tên thành phố trong dữ liệu.")
                
                map_df = map_df.dropna(subset=['lat', 'lon'])
                
                with cm1:
                    if not map_df.empty:
                        # Rename for Pydeck
                        map_df_clean = map_df.rename(columns={COL_CITY: 'City_Name', COL_TOTAL_AMOUNT: 'Total_Cost', COL_BOOKING_CODE: 'Trip_Count'})
                        
                        # Use ScatterplotLayer for Bubble Chart effect
                        # Size scale needs to be adjusted based on data range
                        max_cost = map_df_clean['Total_Cost'].max()
                        map_df_clean['Radius'] = (map_df_clean['Total_Cost'] / max_cost) * 20000 + 5000 # Base 5km + proportional up to 20km
                        
                        layer = pdk.Layer(
                            "ScatterplotLayer",
                            data=map_df_clean,
                            get_position=['lon', 'lat'],
                            get_radius='Radius',
                            get_fill_color=[255, 69, 0, 150], # Orange-Red with transparency
                            pickable=True,
                            auto_highlight=True,
                            stroked=True,
                            filled=True,
                            radius_scale=1,
                            radius_min_pixels=10,
                            radius_max_pixels=100,
                        )
                        
                        view_state = pdk.ViewState(latitude=16.0, longitude=106.0, zoom=4.5, pitch=0) # Flat view for bubbles

                        tooltip = {
                            "html": "<b>{City_Name}</b><br/>Chi phí: {Total_Cost} đ<br/>Số chuyến: {Trip_Count}",
                            "style": {"backgroundColor": "steelblue", "color": "white"}
                        }
                        
                        r = pdk.Deck(
                            layers=[layer],
                            initial_view_state=view_state,
                            tooltip=tooltip,
                            map_style="mapbox://styles/mapbox/light-v9"
                        )
                        st.pydeck_chart(r)
                        st.caption("*Bản đồ hiển thị bong bóng (Bubble Chart), kích thước tương ứng với tổng chi phí.")
                    else:
                        st.info("Chưa có dữ liệu địa điểm hợp lệ.")

                with cm2:
                    st.markdown("##### 📍 Thống Kê Theo Tỉnh/Thành")
                    if COL_CITY in df.columns:
                        pivot_city = df.groupby(COL_CITY).agg({
                            COL_BOOKING_CODE: 'count',
                            COL_TOTAL_AMOUNT: 'sum'
                        }).rename(columns={
                            COL_BOOKING_CODE: 'Số Chuyến', 
                            COL_TOTAL_AMOUNT: 'Tổng Chi'
                        }).sort_values('Số Chuyến', ascending=False)
                        
                        st.dataframe(pivot_city, use_container_width=True)

            # 7. Future Input
            st.markdown("---")
            with st.expander("📥 So Sánh Với Dữ Liệu Cũ (Tính năng đang phát triển)", expanded=False):
                st.info("Tính năng này cho phép bạn tải lên các báo cáo tháng trước để so sánh xu hướng.")
                st.file_uploader("Tải lên file báo cáo cũ (.xlsx)", accept_multiple_files=True, key="old_data_upload")
                st.bar_chart({"Tháng trước": [100, 200, 150], "Tháng này": [120, 180, 210]})

        else:
            st.info("👋 Vui lòng tải file và chạy xử lý dữ liệu ở Tab 2 trước.")    # --- TAB 6: EMAIL TOOL ---
    with tab_email:
        st.header("✉️ Công Cụ Tạo Email & Báo Cáo Gửi Khách Hàng")
        st.markdown("Tạo email thông báo thanh toán Grab tự động dựa trên template và dữ liệu đã xử lý.")
        
        # New Layout: Charts on LEFT, Editor/Actions on RIGHT
        col_email_left, col_email_right = st.columns([1.2, 1])
        
        # --- LEFT COLUMN: CHARTS & STATS ---
        with col_email_left:
            st.subheader("📊 Thống Kê & Trạng Thái")
            
            if 'df_processed' in st.session_state:
                df_proc = st.session_state['df_processed']
                
                if COL_GROUP_FUNCTION in df_proc.columns:
                    # Prepare Summary Data
                    all_funcs = sorted(df_proc[COL_GROUP_FUNCTION].dropna().unique())
                    summary_data = []
                    
                    for func in all_funcs:
                        df_f = df_proc[df_proc[COL_GROUP_FUNCTION] == func]
                        
                        # Count Invoices
                        inv_cnt = 0
                        if COL_BOOKING_CODE in df_f.columns:
                             inv_raw = df_f[COL_BOOKING_CODE].dropna().astype(str).str.strip()
                             inv_cnt = inv_raw[inv_raw != ''].nunique()
                        
                        # Total Amount
                        total_amt = df_f[COL_TOTAL_AMOUNT].sum() if COL_TOTAL_AMOUNT in df_f.columns else 0
                        
                        summary_data.append({
                            'Group Function': func,
                            'Số Hóa Đơn': inv_cnt,
                            'Tổng Tiền': total_amt
                        })
                    
                    df_summary = pd.DataFrame(summary_data)
                    
                    # 1. Bar Chart: Cost by Function
                    fig_cost = px.bar(
                        df_summary.sort_values('Tổng Tiền', ascending=True),
                        x='Tổng Tiền', y='Group Function',
                        orientation='h',
                        text_auto='.2s',
                        title="Top Chi Phí theo Group Function",
                        color='Tổng Tiền',
                        color_continuous_scale='Viridis'
                    )
                    st.plotly_chart(fig_cost, use_container_width=True)
                    
                    # 2. Table Summary
                    st.dataframe(
                        df_summary.style.format({'Tổng Tiền': '{:,.0f}'}), 
                        use_container_width=True,
                        height=400
                    )
                else:
                    st.warning(f"Chưa có cột {COL_GROUP_FUNCTION} để vẽ biểu đồ.")
            else:
                st.info("Vui lòng xử lý dữ liệu trước.")

        # --- RIGHT COLUMN: EDITOR & DOWNLOAD ---
        with col_email_right:
            st.subheader("📝 Template & Tải Về")
            
            # Template Editor
            with st.expander("Chỉnh sửa Template HTML", expanded=False):
                # Load default template content
                default_template = ""
                try:
                    with open("templates/grab_invoice_email.html", "r", encoding="utf-8") as f:
                        default_template = f.read()
                except:
                    default_template = "Template not found."
                    
                # Template Editor
                email_template_content = st.text_area("HTML Template:", value=default_template, height=300)
                
                if st.button("💾 Lưu Template"):
                    try:
                        os.makedirs("templates", exist_ok=True)
                        with open("templates/grab_invoice_email.html", "w", encoding="utf-8") as f:
                            f.write(email_template_content)
                        st.success("✅ Đã lưu template!")
                    except Exception as e:
                        st.error(f"Lỗi khi lưu template: {e}")

            st.markdown("---")
            st.subheader("🚀 Tạo & Tải Email")
            
            if 'df_processed' in st.session_state and up_function:
                # Group Selection for Preview
                df_proc_prev = st.session_state['df_processed']
                all_groups_prev = sorted(df_proc_prev[COL_GROUP_FUNCTION].dropna().unique())
                
                selected_grp = st.selectbox("Chọn nhóm để xem trước:", all_groups_prev)
                
                if selected_grp:
                    # Render Preview for Selected Group
                    try:
                        g_data = st.session_state['df_processed'][st.session_state['df_processed'][COL_GROUP_FUNCTION] == selected_grp]
                        
                        # 1. Lookup Email Info
                        email_map_prev = get_email_mapping_from_upload(up_function)
                        
                        assoc_f = ""
                        if COL_GROUP_FUNCTION in g_data.columns:
                             v_funcs = g_data[COL_GROUP_FUNCTION].dropna()
                             if not v_funcs.empty: assoc_f = v_funcs.iloc[0]
                        
                        l_key = normalize_func_name(str(assoc_f))
                        e_info = email_map_prev.get(l_key, {})
                        
                        if not e_info.get('to'):
                            l_key_d = normalize_func_name(str(selected_grp))
                            e_info_d = email_map_prev.get(l_key_d, {})
                            if e_info_d.get('to'): e_info = e_info_d
                        
                        e_to = e_info.get('to') or "user@example.com"
                        e_cc = e_info.get('cc', "")
                        r_name = extract_name_from_email(e_to)
                        
                        # 2. Stats
                        inv_c = 0
                        if COL_BOOKING_CODE in g_data.columns:
                            inv_raw = g_data[COL_BOOKING_CODE].dropna().astype(str).str.strip()
                            inv_c = inv_raw[inv_raw != ''].nunique()
                        
                        # 3. Date
                        try:
                            ds = pd.to_datetime(g_data[COL_TIME], errors='coerce')
                            my = f"{int(ds.dt.month.max())}/{int(ds.dt.year.max())}"
                        except: my = "MM/YYYY"
                        
                        # 4. Generate HTML
                        preview_html = email_template_content
                        rep_map = {
                            'recipient_name': r_name,
                            'sender_name': 'Duyên',
                            'month_year': my,
                            'invoice_count': inv_c,
                            'vendor_code': 'VENDOR_CODE',
                            'po_number': 'PO_NUMBER',
                            'invoice_date': 'DD-MM-YYYY',
                            'cc_list': e_cc
                        }
                        for k, v in rep_map.items():
                            preview_html = preview_html.replace(f"{{{k}}}", str(v))
                        
                        st.caption(f"📧 Gửi đến: **{e_to}** | CC: {e_cc}")
                        components.html(preview_html, height=500, scrolling=True)
                        
                    except Exception as e:
                        st.error(f"Lỗi xem trước: {e}")
                
                # ACTION BUTTONS
                # Generate All Logic
                if st.button("⚡ Tạo Tất Cả Email (HTML)", type="primary", use_container_width=True):
                    with st.spinner("Đang tạo..."):
                        with tempfile.TemporaryDirectory() as temp_email_dir:
                            email_map_gen = get_email_mapping_from_upload(up_function)
                            
                            cnt = 0
                            for grp in all_groups_prev:
                                try:
                                    # Filter
                                    g_df = df_proc_prev[df_proc_prev[COL_GROUP_FUNCTION] == grp]
                                    
                                    # Email Lookup
                                    assoc_f = ""
                                    if COL_GROUP_FUNCTION in g_df.columns:
                                         v_funcs = g_df[COL_GROUP_FUNCTION].dropna()
                                         if not v_funcs.empty: assoc_f = v_funcs.iloc[0]

                                    l_key = normalize_func_name(str(assoc_f))
                                    e_info = email_map_gen.get(l_key, {})
                                    
                                    # Fallback
                                    if not e_info.get('to'):
                                        l_key_d = normalize_func_name(str(grp))
                                        e_info_d = email_map_gen.get(l_key_d, {})
                                        if e_info_d.get('to'): e_info = e_info_d
                                    
                                    e_to = e_info.get('to') or "user@example.com"
                                    e_cc = e_info.get('cc', "")
                                    r_name = extract_name_from_email(e_to)
                                    
                                    # Stats
                                    inv_c = 0
                                    if COL_BOOKING_CODE in g_df.columns:
                                        inv_raw = g_df[COL_BOOKING_CODE].dropna().astype(str).str.strip()
                                        inv_c = inv_raw[inv_raw != ''].nunique()
                                    
                                    if inv_c == 0: continue
                                    
                                    # Date
                                    try:
                                        ds = pd.to_datetime(g_df[COL_TIME], errors='coerce')
                                        my = f"{int(ds.dt.month.max())}/{int(ds.dt.year.max())}"
                                    except: my = ""
                                    
                                    # Render
                                    final_h = email_template_content
                                    rep_map = {
                                        'recipient_name': r_name,
                                        'sender_name': 'Duyên',
                                        'month_year': my,
                                        'invoice_count': inv_c,
                                        'vendor_code': '94001511',
                                        'po_number': 'PO4502120145',
                                        'invoice_date': '07-08-2023',
                                        'cc_list': e_cc
                                    }
                                    for k, v in rep_map.items():
                                        final_h = final_h.replace(f"{{{k}}}", str(v))
                                        
                                    # Save
                                    safe_n = str(grp).replace('/', '_')
                                    safe_e = e_to.replace('/', '_').replace('\\', '_')
                                    out_name = f"email_{safe_n}_{safe_e}.html"
                                    
                                    with open(os.path.join(temp_email_dir, out_name), "w", encoding="utf-8") as f:
                                        f.write(final_h)
                                    cnt += 1
                                except: pass
                            
                            # Zip result
                            if cnt > 0:
                                zip_buf = io.BytesIO()
                                with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
                                    for r, d, f in os.walk(temp_email_dir):
                                        for file in f:
                                            zf.write(os.path.join(r, file), file)
                                zip_buf.seek(0)
                                st.session_state['email_zip'] = zip_buf
                                st.success(f"Đã tạo {cnt} file.")
                            else:
                                st.warning("Không tạo được file nào.")

                if 'email_zip' in st.session_state:
                    st.download_button(
                        label="⬇️ Tải Xuống Tất Cả Email (Zip)",
                        data=st.session_state['email_zip'],
                        file_name="All_Emails_HTML.zip",
                        mime="application/zip",
                        use_container_width=True
                    )
            
            # --- 3. SUMMARY & CHART (NEW SECTION) ---
            st.markdown("---")
            with st.expander("📊 Bảng Đối Chiếu Số Liệu & Trạng Thái PDF (Theo Group Function)", expanded=True):
                if 'df_processed' in st.session_state:
                    df_proc = st.session_state['df_processed']
                    
                    # Check if COL_GROUP_FUNCTION exists
                    if COL_GROUP_FUNCTION not in df_proc.columns:
                        st.error(f"Không tìm thấy cột '{COL_GROUP_FUNCTION}' trong dữ liệu. Vui lòng kiểm tra cấu hình.")
                    else:
                        all_funcs = sorted(df_proc[COL_GROUP_FUNCTION].dropna().unique())
                        
                        # Get PDF Directory (from session or default)
                        pdf_src_dir = st.session_state.get('src_dir', "./00_master_data/PDF/")
                        
                        pdf_files = []
                        if os.path.exists(pdf_src_dir):
                            try:
                                pdf_files = [f for f in os.listdir(pdf_src_dir) if f.lower().endswith('.pdf')]
                            except: pass
                        
                        summary_data = []
                        
                        # Iterate Group Functions
                        for func in all_funcs:
                            # Filter by Group Function
                            g_data = df_proc[df_proc[COL_GROUP_FUNCTION] == func]
                            
                            # Count from Data
                            count_excel = 0
                            inv_list = []
                            if COL_BOOKING_CODE in g_data.columns:
                                inv_list = g_data[COL_BOOKING_CODE].dropna().astype(str).str.strip().unique()
                                inv_list = [i for i in inv_list if i != '']
                                count_excel = len(inv_list)
                            
                            # Count matching PDFs
                            count_pdf = 0
                            missing_pdf_list = []
                            
                            if pdf_files:
                                # Use Invoice Number for matching if available, else Booking Code
                                match_col = COL_INVOICE_NUM if COL_INVOICE_NUM in g_data.columns else COL_BOOKING_CODE
                                
                                check_items = g_data[match_col].dropna().astype(str).str.strip().unique()
                                check_items = [i for i in check_items if i != '']
                                
                                for item in check_items:
                                    # Naive check: item in filename
                                    if any(item in f for f in pdf_files):
                                        count_pdf += 1
                                    else:
                                        missing_pdf_list.append(item)
                            
                            status = "✅ Đủ" if count_pdf >= count_excel and count_excel > 0 else ("⚠️ Thiếu" if count_pdf > 0 else "❌ Không tìm thấy")
                            if not pdf_files: status = "⚪ Chưa có PDF"
                            
                            summary_data.append({
                                COL_GROUP_FUNCTION: func,
                                "Số HĐ (Excel)": count_excel,
                                "PDF Tìm Thấy": count_pdf if pdf_files else 0,
                                "Trạng Thái": status
                            })
                        
                        df_summary = pd.DataFrame(summary_data)
                        
                        # Display metrics
                        c_sum1, c_sum2 = st.columns(2)
                        if not df_summary.empty:
                            total_excel = df_summary['Số HĐ (Excel)'].sum()
                            total_pdf = df_summary['PDF Tìm Thấy'].sum()
                            
                            c_sum1.metric("Tổng HĐ trên Excel", f"{total_excel:,}")
                            c_sum2.metric("Tổng PDF tìm thấy", f"{total_pdf:,}", delta=f"{total_pdf - total_excel}")
                            
                            # Display Table
                            st.dataframe(df_summary, use_container_width=True)
                            
                            # Chart
                            import plotly.express as px
                            # Melt for grouped bar chart
                            df_melt = df_summary.melt(id_vars=[COL_GROUP_FUNCTION], value_vars=['Số HĐ (Excel)', 'PDF Tìm Thấy'], var_name='Loại', value_name='Số Lượng')
                            
                            fig = px.bar(
                                df_melt, 
                                x=COL_GROUP_FUNCTION, 
                                y='Số Lượng', 
                                color='Loại',
                                barmode='group',
                                title="So Sánh Số Lượng Hóa Đơn: Dữ Liệu Excel vs File PDF",
                                text_auto=True,
                                color_discrete_map={'Số HĐ (Excel)': '#00B14F', 'PDF Tìm Thấy': '#FF6B6B'} # Grab Green vs Red
                            )
                            st.plotly_chart(fig, use_container_width=True)
                            
                            if not pdf_files:
                                st.warning(f"Không tìm thấy file PDF nào trong thư mục: '{pdf_src_dir}'. Vui lòng kiểm tra lại đường dẫn ở Tab 'Phân Phối PDF'.")        

    # --- TAB 7: OPEN EMAIL DRAFT ---
    # --- TAB 7: KIỂM TRA EMAIL DRAFT (LOCAL VIEW ONLY) ---
    with tab_draft:
        st.header("📧 Kiểm Tra & Mở Email Draft (Local Only)")
        st.markdown("""
        **Lưu ý:** Tab này chỉ hoạt động khi chạy trên máy tính cá nhân (Localhost) để kiểm tra kết quả trong thư mục đầu ra cũ. 
        Nếu bạn đang dùng quy trình "Phân Phối & Đóng Gói" (Tab 5), file Email Draft đã có sẵn trong file Zip tải về.
        """)
        
        st.markdown("---")
        
        # Folder Selection Logic
        c1, c2 = st.columns([3, 1])
        
        # Init paths
        if 'dst_dir_draft' not in st.session_state: 
             # Try to pick up from Tab 5 if it was set manually there, else default
             st.session_state['dst_dir_draft'] = st.session_state.get('dst_dir', "./01_function_BK/Invoices_by_Function/")

        def update_draft_path():
            st.session_state['dst_dir_draft'] = st.session_state.input_draft_path

        with c1:
             dst_dir_draft_input = st.text_input(
                 "Thư mục chứa hồ sơ (Đã phân phối):", 
                 value=st.session_state['dst_dir_draft'],
                 key="input_draft_path",
                 on_change=update_draft_path
             )
        
        # Ensure variable is sync
        dst_dir_draft = st.session_state['dst_dir_draft']
        
        with c2:
             if st.button("📂 Chọn Thư Mục", key="btn_choose_draft_folder"):
                 import tkinter as tk
                 from tkinter import filedialog
                 try:
                     root = tk.Tk()
                     root.withdraw()
                     root.wm_attributes('-topmost', 1)
                     selected = filedialog.askdirectory(master=root)
                     root.destroy()
                     if selected:
                         st.session_state['dst_dir_draft'] = selected
                         st.session_state['input_draft_path'] = selected
                         st.rerun()
                 except: pass

        # --- PREVIEW SECTION ---
        st.markdown("### 👁️ Danh Sách Nhóm & Mở Draft")
        
        if st.button("🔍 Quét Thư Mục & Tải Lại"):
             st.rerun()

        if os.path.exists(dst_dir_draft):
            try:
                found_groups = [d for d in os.listdir(dst_dir_draft) if os.path.isdir(os.path.join(dst_dir_draft, d))]
                
                if not found_groups:
                     st.warning("Không tìm thấy thư mục con nào trong đường dẫn này.")
                else:
                    st.success(f"Tìm thấy {len(found_groups)} nhóm.")
                    
                    selected_preview_group = st.selectbox("Chọn nhóm để xử lý:", found_groups)
                    
                    if selected_preview_group:
                        g_path = os.path.join(dst_dir_draft, selected_preview_group)
                        g_files = os.listdir(g_path)
                        
                        # Find key files
                        html_f_name = None
                        # Try exact group match first
                        if f"{selected_preview_group}.html" in g_files:
                             html_f_name = f"{selected_preview_group}.html"
                        # Try prefix match (old logic)
                        elif any(f.startswith(f"email_{selected_preview_group}") and f.endswith(".html") for f in g_files):
                             html_f_name = next(f for f in g_files if f.startswith(f"email_{selected_preview_group}") and f.endswith(".html"))
                        # Fallback any html
                        elif any(f.endswith(".html") for f in g_files):
                             html_f_name = next(f for f in g_files if f.endswith(".html"))

                        html_exists = html_f_name is not None
                        
                        excel_f = next((f for f in g_files if f.endswith('.xlsx') and f.startswith('BK_GRAB_')), None)
                        
                        # Count PDF recursively
                        pdf_count = 0
                        for root, dirs, files in os.walk(g_path):
                            for f in files:
                                if f.lower().endswith('.pdf'): pdf_count += 1
                        
                        # Get Email info
                        current_email = "Không tìm thấy"
                        if html_f_name:
                            if "email_" in html_f_name:
                                 current_email = html_f_name.replace(f"email_{selected_preview_group}_", "").replace(".html", "")
                            else:
                                 current_email = "(Xem trong file HTML)"

                        # Display Info
                        st.info(f"**Email:** {current_email} | **Excel:** {excel_f if excel_f else '❌'} | **PDF:** {pdf_count}")
                        
                        st.markdown("---")
                        
                        # OPEN DRAFT BUTTON
                        col_btn1, col_btn2 = st.columns(2)
                        
                        with col_btn1:
                            if st.button("🚀 Tạo & Mở File Draft (Local)"):
                                if not html_exists or not excel_f:
                                    st.error("Thiếu file HTML hoặc Excel để tạo Draft.")
                                else:
                                    # Standalone Single-File Draft Creation Logic
                                    try:
                                        # 1. Read HTML
                                        with open(os.path.join(g_path, html_f_name), 'r', encoding='utf-8') as f:
                                            email_body = f.read()

                                        msg = MIMEMultipart('mixed')
                                        msg['Subject'] = f"Bảng Kê Grab Tháng - Nhóm {selected_preview_group}"
                                        
                                        if "@" in current_email: msg['To'] = current_email
                                        else: msg['To'] = ""
                                        msg['X-Unsent'] = '1'

                                        # Body
                                        msg_body = MIMEMultipart('alternative')
                                        msg_body.attach(MIMEText("Vui lòng xem ở chế độ HTML.", 'plain', 'utf-8'))
                                        msg_body.attach(MIMEText(email_body, 'html', 'utf-8'))
                                        msg.attach(msg_body)

                                        # Excel
                                        excel_path = os.path.join(g_path, excel_f)
                                        with open(excel_path, "rb") as f:
                                            part = MIMEApplication(f.read(), Name=excel_f)
                                        part['Content-Disposition'] = f'attachment; filename="{excel_f}"'
                                        msg.attach(part)

                                        # Zip PDFs
                                        zip_filename = f"Invoices_{selected_preview_group}.zip"
                                        zip_path = os.path.join(g_path, zip_filename)
                                        
                                        if os.path.exists(zip_path): 
                                            try: os.remove(zip_path) 
                                            except: pass

                                        pdf_files_to_zip = []
                                        for root, dirs, files_in_dir in os.walk(g_path):
                                            for file in files_in_dir:
                                                if file.lower().endswith('.pdf'):
                                                    pdf_files_to_zip.append(os.path.join(root, file))
                                        
                                        if pdf_files_to_zip:
                                            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                                                for pdf_p in pdf_files_to_zip:
                                                    rel_path = os.path.relpath(pdf_p, g_path)
                                                    zipf.write(pdf_p, arcname=rel_path)
                                            
                                            with open(zip_path, 'rb') as f:
                                                part = MIMEApplication(f.read(), Name=zip_filename)
                                            part['Content-Disposition'] = f'attachment; filename="{zip_filename}"'
                                            msg.attach(part)

                                        # Save EML
                                        eml_name = f"Draft_{selected_preview_group}.eml"
                                        eml_path = os.path.join(g_path, eml_name)
                                        with open(eml_path, 'wb') as f:
                                            f.write(msg.as_bytes())

                                        # Open Folder
                                        subprocess.Popen(f'explorer /select,"{os.path.abspath(eml_path)}"')
                                        st.success(f"✅ Đã tạo & mở: {eml_name}")

                                    except Exception as e:
                                        st.error(f"Lỗi: {e}")

                        with col_btn2:
                             if st.button("📂 Mở Thư Mục Nhóm"):
                                 try: os.startfile(g_path)
                                 except: st.error("Không thể mở thư mục.")

            except Exception as e:
                st.error(f"Lỗi đọc thư mục: {e}")
        else:
            st.error(f"Thư mục không tồn tại: {dst_dir_draft}")

